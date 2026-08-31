from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DISCOURSE_XLSX = BASE_DIR / "Prabowo Speeches - Discourse Situational Context Analysis.xlsx"
VISUAL_XLSX = BASE_DIR / "Prabowo Speeches - Visual Mapping Counts.xlsx"
OUT_DIR = BASE_DIR / "interactive_dashboard"
OUT_FILE = OUT_DIR / "index.html"
MOBILE_OUT_DIR = BASE_DIR / "mobile_dashboard"
MOBILE_OUT_FILE = MOBILE_OUT_DIR / "index.html"
ANALYSIS_DIR = BASE_DIR / "analysis"

ANALYSIS_TABLE_TITLES = {
    "T0_per_speech_master": "Per-Speech Master",
    "T1_selfmention_full_corpus": "Self-Mention: Full Corpus",
    "T1b_selfmention_EXCLUSIONS_APPLIED": "Self-Mention: Exclusions Applied",
    "T2_selfmention_sample": "Self-Mention: Purposive Samples",
    "T3_selfmention_by_field_domain": "Self-Mention by Field Domain",
    "T4_selfmention_by_interaction_type": "Self-Mention by Interaction Type",
    "T5_appraisal_attitude_by_pronoun": "Appraisal Attitude by Pronoun",
    "T5b_attitude_group_summary": "Appraisal Attitude Summary",
    "T6_engagement_hedge_booster": "Engagement: Hedges and Boosters",
    "T7_firstperson_verb_frames": "First-Person Verb Frames",
    "T8_data_hygiene_log": "Data Hygiene Log",
    "T9_speech_to_table_map": "Speech-to-Table Audit Map",
}


def cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value


def csv_value(header: str, value: str) -> Any:
    value = value.strip()
    if not value:
        return ""
    if header == "id":
        return value
    if value in {"True", "False"}:
        return value == "True"
    try:
        return float(value) if "." in value else int(value)
    except ValueError:
        return value


def csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [
            {header: csv_value(header, value or "") for header, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def research_analysis() -> dict[str, Any]:
    tables = {}
    for path in sorted(ANALYSIS_DIR.glob("T*.csv")):
        key = path.stem
        tables[key] = {
            "title": ANALYSIS_TABLE_TITLES.get(key, key.replace("_", " ")),
            "source": path.name,
            "rows": csv_rows(path),
        }
    sample_path = ANALYSIS_DIR / "sample_definition.json"
    samples = json.loads(sample_path.read_text(encoding="utf-8")) if sample_path.exists() else {}
    return {"tables": tables, "samples": samples}


def sheet_rows(path: Path, sheet_name: str, header_row: int = 1) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers = [cell_value(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = {}
        for header, value in zip(headers, row):
            if header:
                record[str(header)] = cell_value(value)
        if any(value is not None and value != "" for value in record.values()):
            rows.append(record)
    return rows


def framework_rows() -> list[dict[str, Any]]:
    wb = load_workbook(DISCOURSE_XLSX, data_only=True, read_only=True)
    ws = wb["Framework"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=3, max_row=8, values_only=True):
        if not row[0]:
            continue
        rows.append(
            {
                "term": cell_value(row[0]),
                "meaning": cell_value(row[1]),
                "components": cell_value(row[2]),
            }
        )
    return rows


def summary_rows() -> list[dict[str, Any]]:
    wb = load_workbook(DISCOURSE_XLSX, data_only=True, read_only=True)
    ws = wb["Summary"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        metric, category, count, share = row[:4]
        if metric is None and category is None:
            continue
        rows.append(
            {
                "metric": cell_value(metric),
                "category": cell_value(category),
                "count": cell_value(count),
                "share": cell_value(share),
            }
        )
    return rows


def visual_mapping() -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(VISUAL_XLSX, data_only=True, read_only=True)
    ws = wb["Visual Mapping"]
    interaction: list[dict[str, Any]] = []
    domain: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0]:
            interaction.append({"category": cell_value(row[0]), "count": cell_value(row[1]), "share": cell_value(row[2])})
        if row[4]:
            domain.append({"category": cell_value(row[4]), "count": cell_value(row[5]), "share": cell_value(row[6])})
    return {"interactionTypes": interaction, "fieldDomains": domain}


def normalize_analysis(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        normalized.append(
            {
                "no": str(row.get("No", "") or ""),
                "date": str(row.get("Date", "") or ""),
                "year": str(row.get("Year", "") or ""),
                "language": str(row.get("Language Group", "") or ""),
                "event": str(row.get("Speech/Event", "") or ""),
                "interactionType": str(row.get("Interaction Type", "") or ""),
                "fieldDomain": str(row.get("Field Domain", "") or ""),
                "field": str(row.get("Field", "") or ""),
                "tenor": str(row.get("Tenor", "") or ""),
                "mode": str(row.get("Mode", "") or ""),
                "openingEvidence": str(row.get("Opening Evidence", "") or ""),
                "wordCount": int(row.get("Word Count") or 0),
                "filePath": Path(str(row.get("File Path", "") or "")).name,
            }
        )
    return normalized


def build_html(data: dict[str, Any]) -> str:
    data_json = json.dumps(data, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Prabowo Speeches Discourse Dashboard</title>
  <style>
    :root {{
      color-scheme: light;
      --paper:    #f4f1ea;
      --paper-2:  #ebe5d8;
      --panel:    #fbf9f3;
      --ink:      #1b1a16;
      --ink-soft: #3b3731;
      --muted:    #6b6458;
      --faint:    #9a9286;
      --line:     #dcd5c6;
      --line-2:   #c8c0ad;
      --rule:     #26231d;
      --accent:      #9a2b27;
      --accent-deep: #6f1d1a;
      --slate:    #2f4858;
      --ochre:    #876c34;
      --serif: "Iowan Old Style", "Palatino Linotype", Palatino, Charter, Georgia, "Times New Roman", serif;
      --sans: -apple-system, BlinkMacSystemFont, "Helvetica Neue", "Segoe UI", Roboto, Arial, sans-serif;
      --mono: "SF Mono", ui-monospace, "DejaVu Sans Mono", "Roboto Mono", monospace;
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      font-family: var(--sans);
      font-size: 15px;
      background: var(--paper);
      color: var(--ink);
      -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility;
    }}

    ::selection {{ background: rgba(154, 43, 39, 0.16); }}

    /* Masthead ------------------------------------------------------------ */
    header {{
      background: var(--paper);
      border-top: 3px solid var(--accent);
    }}

    .shell {{
      width: min(1280px, calc(100vw - 56px));
      margin: 0 auto;
    }}

    .masthead {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 48px;
      padding: 38px 0 26px;
      align-items: end;
    }}

    .hero-kicker {{
      margin: 0 0 18px;
      color: var(--accent);
      font-size: 12px;
      font-weight: 700;
      line-height: 1;
      text-transform: uppercase;
      letter-spacing: 0.2em;
    }}

    h1 {{
      margin: 0;
      max-width: 17ch;
      font-family: var(--serif);
      font-size: clamp(38px, 5vw, 62px);
      font-weight: 600;
      line-height: 1.02;
      letter-spacing: -0.015em;
      color: var(--ink);
    }}

    .subtitle {{
      margin: 18px 0 0;
      max-width: 62ch;
      color: var(--muted);
      font-family: var(--serif);
      font-size: 18px;
      line-height: 1.5;
    }}

    .source-stack {{
      display: flex;
      flex-direction: column;
      gap: 0;
      align-items: flex-end;
      align-self: center;
      max-width: 300px;
    }}

    .pill {{
      display: block;
      width: 100%;
      padding: 9px 0;
      border-top: 1px solid var(--line);
      color: var(--muted);
      font-size: 11px;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.07em;
      text-align: right;
      line-height: 1.4;
    }}

    /* Tabs ---------------------------------------------------------------- */
    .tabs {{
      display: flex;
      gap: 30px;
      overflow-x: auto;
      padding: 0;
      margin-top: 10px;
      border-bottom: 1px solid var(--rule);
    }}

    .tab-btn, button, select, input {{ font: inherit; }}

    .tab-btn {{
      border: 0;
      background: transparent;
      color: var(--muted);
      padding: 13px 0;
      margin-bottom: -1px;
      border-bottom: 2px solid transparent;
      cursor: pointer;
      font-size: 14px;
      font-weight: 600;
      letter-spacing: 0.01em;
      white-space: nowrap;
      min-width: max-content;
    }}

    .tab-btn:hover {{ color: var(--ink); }}

    .tab-btn.active {{
      color: var(--ink);
      border-bottom-color: var(--accent);
    }}

    main {{ padding: 28px 0 56px; }}

    /* Filters ------------------------------------------------------------- */
    .filters {{
      position: sticky;
      top: 0;
      z-index: 5;
      background: var(--paper);
      border-bottom: 1px solid var(--line);
      padding: 14px 0;
      margin-bottom: 28px;
    }}

    .filters.is-hidden {{ display: none; }}

    .filter-grid {{
      display: grid;
      grid-template-columns: 1.4fr repeat(4, minmax(140px, 1fr)) auto;
      gap: 10px;
      align-items: center;
    }}

    input, select {{
      width: 100%;
      height: 40px;
      border: 1px solid var(--line-2);
      border-radius: 2px;
      background: var(--panel);
      color: var(--ink);
      padding: 0 11px;
      font-family: var(--sans);
      font-size: 14px;
      outline: none;
    }}

    input::placeholder {{ color: var(--faint); }}

    input:focus, select:focus {{
      border-color: var(--accent);
      box-shadow: inset 0 0 0 1px var(--accent);
    }}

    .clear-btn {{
      height: 40px;
      border: 1px solid var(--rule);
      color: var(--ink);
      background: transparent;
      border-radius: 2px;
      padding: 0 18px;
      cursor: pointer;
      font-size: 12px;
      font-weight: 600;
      letter-spacing: 0.05em;
      text-transform: uppercase;
      white-space: nowrap;
    }}

    .clear-btn:hover {{ background: var(--ink); color: var(--paper); border-color: var(--ink); }}

    .view {{ display: none; }}
    .view.active {{ display: block; }}

    /* KPI strip ----------------------------------------------------------- */
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(5, minmax(0, 1fr));
      gap: 0;
      margin-bottom: 36px;
      border-top: 2px solid var(--rule);
    }}

    .kpi {{
      padding: 18px 24px 8px;
      display: flex;
      flex-direction: column;
      gap: 12px;
      border-right: 1px solid var(--line);
    }}

    .kpi:first-child {{ padding-left: 0; }}
    .kpi:last-child {{ border-right: 0; padding-right: 0; }}

    .kpi-label {{
      color: var(--muted);
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: 0.12em;
      font-weight: 700;
    }}

    .kpi-value {{
      font-family: var(--serif);
      font-size: clamp(28px, 3vw, 40px);
      line-height: 1;
      font-weight: 600;
      color: var(--ink);
      font-variant-numeric: tabular-nums;
    }}

    .kpi-note {{
      color: var(--muted);
      font-size: 12.5px;
      line-height: 1.4;
    }}

    /* Layout -------------------------------------------------------------- */
    .grid-2 {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      gap: 24px;
    }}

    .grid-3 {{
      display: grid;
      grid-template-columns: 1.1fr 0.9fr 1fr;
      gap: 24px;
    }}

    .full-width-panel {{ margin-top: 24px; }}

    .panel {{
      padding: 24px;
      min-width: 0;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 3px;
    }}

    .panel h2, .panel h3 {{
      margin: 0 0 16px;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
      font-family: var(--serif);
      font-size: 19px;
      font-weight: 600;
      line-height: 1.2;
      letter-spacing: -0.005em;
      color: var(--ink);
    }}

    .panel-caption {{
      color: var(--muted);
      font-size: 13px;
      margin: -8px 0 16px;
      line-height: 1.45;
    }}

    .chart {{ min-height: 240px; }}

    /* Bars ---------------------------------------------------------------- */
    .bar-row {{
      display: grid;
      grid-template-columns: minmax(160px, 1fr) minmax(120px, 2fr) 96px;
      gap: 14px;
      align-items: center;
      margin: 13px 0;
      font-size: 13px;
    }}

    .bar-label {{
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
      color: var(--ink-soft);
    }}

    .bar-track {{
      height: 10px;
      border-radius: 2px;
      background: var(--paper-2);
      overflow: hidden;
    }}

    .bar-fill {{
      height: 100%;
      border-radius: 2px;
      background: var(--accent);
    }}

    .bar-fill.coral {{ background: var(--accent); }}
    .bar-fill.indigo {{ background: var(--slate); }}

    .bar-value {{
      text-align: right;
      color: var(--muted);
      font-family: var(--mono);
      font-size: 12px;
      font-variant-numeric: tabular-nums;
    }}

    .bar-value span {{ color: var(--faint); }}

    /* Donut --------------------------------------------------------------- */
    .donut-wrap {{
      display: grid;
      grid-template-columns: 168px minmax(0, 1fr);
      gap: 30px;
      align-items: center;
      min-height: 220px;
    }}

    .donut {{
      width: 168px;
      aspect-ratio: 1;
      border-radius: 50%;
      position: relative;
    }}

    .donut::after {{
      content: "";
      position: absolute;
      inset: 32px;
      border-radius: 50%;
      background: var(--panel);
    }}

    .legend {{ display: grid; gap: 0; }}

    .legend-item {{
      display: grid;
      grid-template-columns: 10px minmax(0, 1fr) auto;
      gap: 12px;
      align-items: center;
      font-size: 13px;
      padding: 11px 0;
      border-bottom: 1px solid var(--line);
    }}

    .legend-item:last-child {{ border-bottom: 0; }}

    .legend-item strong {{
      font-family: var(--mono);
      font-weight: 600;
      font-variant-numeric: tabular-nums;
      color: var(--ink);
    }}

    .swatch {{
      width: 10px;
      height: 10px;
      border-radius: 0;
      background: var(--accent);
    }}

    .swatch.coral {{ background: var(--accent); }}
    .swatch.indigo {{ background: var(--slate); }}
    .swatch.amber {{ background: var(--ochre); }}

    /* Timeline ------------------------------------------------------------ */
    .timeline {{ min-height: 240px; padding-top: 6px; }}

    .line-chart {{ width: 100%; height: 248px; display: block; }}
    .line-axis {{ stroke: var(--line-2); stroke-width: 1; }}
    .line-grid {{ stroke: var(--line); stroke-width: 1; stroke-dasharray: 2 3; }}
    .line-path {{
      fill: none;
      stroke: var(--accent);
      stroke-width: 2;
      stroke-linejoin: round;
    }}
    .line-area {{ fill: rgba(154, 43, 39, 0.07); }}
    .line-point {{
      fill: var(--panel);
      stroke: var(--accent);
      stroke-width: 2;
    }}
    .line-label {{ fill: var(--muted); font-family: var(--mono); font-size: 11px; }}
    .line-value {{ fill: var(--ink); font-family: var(--mono); font-size: 11px; font-weight: 600; }}

    /* Heatmap ------------------------------------------------------------- */
    .heatmap {{
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: var(--panel);
    }}

    .heatmap-table {{ font-family: var(--sans); }}

    .heatmap-table th {{
      min-width: 116px;
      color: var(--ink);
      font-size: 11px;
      font-weight: 700;
      line-height: 1.3;
      letter-spacing: 0.04em;
      text-transform: uppercase;
      cursor: default;
      background: var(--paper-2);
    }}

    .heatmap-table td {{
      font-size: 13px;
      line-height: 1.35;
      font-family: var(--mono);
    }}

    .heatmap-table td:first-child {{
      min-width: 168px;
      font-family: var(--sans);
    }}

    /* Tables -------------------------------------------------------------- */
    table {{ width: 100%; border-collapse: collapse; }}

    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 11px 12px;
      text-align: left;
      vertical-align: top;
      font-size: 13px;
      line-height: 1.4;
    }}

    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: var(--paper-2);
      color: var(--ink);
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      cursor: pointer;
      user-select: none;
      border-bottom: 2px solid var(--rule);
    }}

    tbody tr:hover {{ background: var(--paper-2); }}

    .table-wrap {{
      max-height: 680px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: var(--panel);
    }}

    /* Framework ----------------------------------------------------------- */
    .framework-grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 18px;
    }}

    .framework-item {{
      padding: 20px;
      cursor: pointer;
      min-height: 168px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-left: 3px solid var(--line-2);
      border-radius: 3px;
      transition: border-color 120ms ease;
    }}

    .framework-item:hover {{ border-left-color: var(--accent); }}

    .framework-item.active {{
      border-color: var(--line);
      border-left: 3px solid var(--accent);
    }}

    .framework-item h3 {{
      margin: 0 0 12px;
      font-family: var(--serif);
      font-size: 22px;
      font-weight: 600;
      color: var(--accent);
    }}

    .framework-item p {{
      margin: 0;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.5;
    }}

    .framework-item p strong {{ color: var(--ink-soft); }}

    .detail-box {{
      margin-top: 20px;
      padding: 22px 24px;
      border: 1px solid var(--line);
      border-left: 3px solid var(--accent);
      background: var(--panel);
      border-radius: 3px;
    }}

    .detail-box h3 {{
      margin: 0 0 10px;
      font-family: var(--serif);
      font-size: 20px;
      font-weight: 600;
      color: var(--ink);
    }}

    .detail-box p {{
      margin: 0 0 8px;
      color: var(--ink-soft);
      font-size: 14px;
      line-height: 1.55;
    }}

    .detail-box p strong {{ color: var(--ink); }}

    /* Summary ------------------------------------------------------------- */
    .summary-list {{ display: grid; gap: 10px; }}

    .summary-cards {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 24px;
    }}

    .summary-table th {{
      position: static;
      cursor: pointer;
      border-bottom: 1px solid var(--rule);
    }}

    .summary-table th.active-sort {{ color: var(--accent); }}

    .summary-table td {{ font-size: 13px; }}

    .summary-table td:last-child,
    .summary-table th:last-child {{
      text-align: right;
      font-family: var(--mono);
      font-variant-numeric: tabular-nums;
    }}

    .summary-row {{
      display: grid;
      grid-template-columns: 180px minmax(0, 1fr) 90px 90px;
      gap: 12px;
      align-items: center;
      padding: 12px 14px;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: var(--panel);
    }}

    .summary-row strong {{ font-size: 13px; }}
    .summary-row span {{ color: var(--muted); font-size: 13px; }}

    /* Visual mapping ------------------------------------------------------ */
    .visual-map {{ display: grid; gap: 8px; }}

    .visual-map .bar-row {{
      grid-template-columns: minmax(260px, 1.6fr) minmax(150px, 1fr) 124px;
      align-items: center;
    }}

    .visual-map .bar-label {{
      overflow: visible;
      text-overflow: clip;
      white-space: normal;
      line-height: 1.35;
    }}

    .visual-map .bar-value {{ text-align: right; white-space: nowrap; }}

    /* Analysis ------------------------------------------------------------ */
    .analysis-toolbar {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      margin-bottom: 14px;
      color: var(--muted);
      font-size: 13px;
    }}

    .analysis-toolbar span:first-child {{ color: var(--ink); font-weight: 600; }}

    .evidence {{ color: var(--muted); max-width: 540px; line-height: 1.5; }}

    /* Linguistic analysis ------------------------------------------------ */
    .research-head {{
      display: grid;
      grid-template-columns: minmax(0, 1.5fr) minmax(240px, 0.5fr);
      gap: 28px;
      align-items: end;
      padding: 4px 0 24px;
      border-bottom: 2px solid var(--rule);
      margin-bottom: 26px;
    }}

    .research-head h2 {{
      margin: 0 0 8px;
      font-family: var(--serif);
      font-size: 30px;
      font-weight: 600;
    }}

    .research-head p {{ margin: 0; color: var(--muted); line-height: 1.55; max-width: 72ch; }}

    .sample-note {{
      color: var(--ink-soft);
      font-family: var(--mono);
      font-size: 12px;
      line-height: 1.6;
      text-align: right;
    }}

    .research-kpis {{ margin-bottom: 28px; }}

    .research-nav {{
      display: flex;
      gap: 0;
      overflow-x: auto;
      margin: 0 0 22px;
      border: 1px solid var(--line-2);
      background: var(--panel);
    }}

    .research-nav button {{
      min-width: max-content;
      padding: 11px 16px;
      border: 0;
      border-right: 1px solid var(--line);
      background: transparent;
      color: var(--muted);
      cursor: pointer;
      font-size: 12px;
      font-weight: 700;
    }}

    .research-nav button:last-child {{ border-right: 0; }}
    .research-nav button:hover {{ color: var(--ink); background: var(--paper-2); }}
    .research-nav button.active {{ color: var(--panel); background: var(--ink); }}

    .research-panel {{ display: none; }}
    .research-panel.active {{ display: block; }}

    .analysis-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 24px;
    }}

    .analysis-grid .wide {{ grid-column: 1 / -1; }}

    .rate-chart {{ display: grid; gap: 14px; }}

    .rate-row {{
      display: grid;
      grid-template-columns: minmax(130px, 0.9fr) minmax(180px, 2fr) 84px;
      gap: 14px;
      align-items: center;
      font-size: 13px;
    }}

    .rate-label strong {{ display: block; color: var(--ink); }}
    .rate-label span {{ display: block; margin-top: 2px; color: var(--muted); font-size: 11px; }}
    .rate-value {{ text-align: right; font-family: var(--mono); color: var(--ink-soft); }}

    .grouped-track {{
      display: grid;
      gap: 4px;
      min-width: 0;
    }}

    .grouped-track .bar-track {{ height: 8px; }}
    .bar-fill.slate {{ background: var(--slate); }}
    .bar-fill.ochre {{ background: var(--ochre); }}

    .analysis-legend {{
      display: flex;
      flex-wrap: wrap;
      gap: 16px;
      margin-bottom: 16px;
      color: var(--muted);
      font-size: 12px;
    }}

    .analysis-legend span {{ display: inline-flex; gap: 7px; align-items: center; }}

    .attitude-row {{ margin: 17px 0; }}
    .attitude-row-head {{ display: flex; justify-content: space-between; gap: 12px; margin-bottom: 7px; font-size: 13px; }}
    .attitude-row-head span {{ color: var(--muted); font-family: var(--mono); }}

    .stacked-bar {{
      display: flex;
      width: 100%;
      height: 14px;
      overflow: hidden;
      background: var(--paper-2);
    }}

    .stacked-bar i {{ display: block; height: 100%; }}

    .finding {{
      padding: 14px 0;
      border-bottom: 1px solid var(--line);
      color: var(--ink-soft);
      font-size: 13px;
      line-height: 1.5;
    }}

    .finding:last-child {{ border-bottom: 0; }}
    .finding strong {{ color: var(--accent); font-family: var(--serif); font-size: 22px; margin-right: 8px; }}

    .source-toolbar {{
      display: grid;
      grid-template-columns: minmax(240px, 0.7fr) minmax(240px, 1fr) auto;
      gap: 10px;
      align-items: center;
      margin-bottom: 12px;
    }}

    .source-meta {{ color: var(--muted); font-size: 12px; margin: 0 0 14px; }}
    .source-table th {{ white-space: nowrap; }}
    .source-table td {{ max-width: 360px; overflow-wrap: anywhere; }}
    .source-table td.numeric {{ text-align: right; font-family: var(--mono); }}

    .master-table {{
      min-width: 3380px;
      table-layout: fixed;
    }}

    .master-table th,
    .master-table td {{
      max-width: none;
      white-space: nowrap;
    }}

    .master-table .master-groups th {{
      top: 0;
      height: 34px;
      padding: 8px 12px;
      color: var(--panel);
      background: var(--ink);
      border-right: 1px solid var(--ink-soft);
      border-bottom: 0;
      cursor: default;
      text-align: center;
      letter-spacing: 0.09em;
    }}

    .master-table .master-groups th:first-child {{
      position: sticky;
      left: 0;
      z-index: 6;
    }}

    .master-table thead tr:nth-child(2) th {{ top: 34px; }}

    .master-table col:nth-child(1) {{ width: 62px; }}
    .master-table col:nth-child(2) {{ width: 112px; }}
    .master-table col:nth-child(3) {{ width: 150px; }}
    .master-table col:nth-child(4) {{ width: 245px; }}
    .master-table col:nth-child(5) {{ width: 330px; }}
    .master-table col:nth-child(n+6):nth-child(-n+24) {{ width: 103px; }}
    .master-table th:nth-child(5), .master-table td:nth-child(5) {{ white-space: normal; }}
    .master-table th:nth-child(9), .master-table td:nth-child(9),
    .master-table th:nth-child(10), .master-table td:nth-child(10) {{ text-align: center; }}
    .master-table col:nth-child(25) {{ width: 520px; }}
    .master-table th:nth-child(25), .master-table td:nth-child(25) {{ white-space: normal; }}

    .master-table thead tr:nth-child(2) th:nth-child(-n+3),
    .master-table tbody td:nth-child(-n+3) {{
      position: sticky;
      z-index: 2;
      background: var(--panel);
    }}

    .master-table thead tr:nth-child(2) th:nth-child(-n+3) {{ z-index: 5; background: var(--paper-2); }}
    .master-table th:nth-child(1), .master-table td:nth-child(1) {{ left: 0; }}
    .master-table th:nth-child(2), .master-table td:nth-child(2) {{ left: 62px; }}
    .master-table th:nth-child(3), .master-table td:nth-child(3) {{ left: 174px; box-shadow: 6px 0 8px -8px rgba(27, 26, 22, 0.65); }}
    .master-table tbody tr:hover td:nth-child(-n+3) {{ background: var(--paper-2); }}

    .empty {{
      padding: 32px;
      color: var(--muted);
      text-align: center;
      border: 1px dashed var(--line-2);
      border-radius: 3px;
      background: var(--panel);
    }}

    @media (max-width: 1100px) {{
      .masthead, .grid-2, .grid-3, .filter-grid, .summary-cards, .research-head, .source-toolbar {{
        grid-template-columns: 1fr;
      }}
      .source-stack {{ align-items: flex-start; max-width: none; }}
      .pill {{ text-align: left; }}
      .kpi-grid {{
        grid-template-columns: repeat(2, minmax(0, 1fr));
        border-top: 0;
        gap: 0 24px;
      }}
      .kpi {{
        border-right: 0;
        border-top: 2px solid var(--rule);
        padding: 16px 0 8px;
      }}
      .framework-grid {{ grid-template-columns: 1fr; }}
      .donut-wrap {{ grid-template-columns: 1fr; }}
      .sample-note {{ text-align: left; }}
    }}

    @media (max-width: 640px) {{
      .shell {{ width: min(100vw - 28px, 1280px); }}
      .kpi-grid {{ grid-template-columns: 1fr; }}
      .bar-row {{ grid-template-columns: 1fr; gap: 6px; }}
      .bar-value {{ text-align: left; }}
      .summary-row {{ grid-template-columns: 1fr; }}
      .visual-map .bar-row {{ grid-template-columns: 1fr; }}
      .visual-map .bar-value {{ text-align: left; }}
      .analysis-grid {{ grid-template-columns: 1fr; }}
      .analysis-grid .wide {{ grid-column: auto; }}
      .rate-row {{ grid-template-columns: 1fr; gap: 6px; }}
      .rate-value {{ text-align: left; }}
    }}
  
</style>
</head>
<body>
  <header>
    <div class="shell">
      <div class="masthead">
        <div>
          <p class="hero-kicker">Situational Context Analysis</p>
          <h1>Prabowo Speeches Discourse Dashboard</h1>
          <p class="subtitle">Explore situational context across 169 speech transcripts, then move into self-mention, appraisal, engagement, first-person verb framing, and the underlying research tables.</p>
        </div>
        <div class="source-stack">
          <span class="pill">Discourse workbook: Framework, Summary, Analysis Table</span>
          <span class="pill">Visual mapping workbook: Interaction Types and Field Domains</span>
          <span class="pill">Analysis folder: Linguistic patterns and data audit</span>
        </div>
      </div>
      <nav class="tabs" aria-label="Dashboard views">
        <button class="tab-btn active" data-view="overview">Overview</button>
        <button class="tab-btn" data-view="framework">Framework</button>
        <button class="tab-btn" data-view="summary">Summary</button>
        <button class="tab-btn" data-view="analysis">Analysis Table</button>
        <button class="tab-btn" data-view="linguistic">Linguistic Analysis</button>
      </nav>
    </div>
  </header>

  <section class="filters">
    <div class="shell filter-grid">
      <input id="searchInput" type="search" placeholder="Search events, evidence, fields" />
      <select id="languageFilter"></select>
      <select id="yearFilter"></select>
      <select id="interactionFilter"></select>
      <select id="domainFilter"></select>
      <button class="clear-btn" id="clearFilters">Reset</button>
    </div>
  </section>

  <main class="shell">
    <section id="overview" class="view active">
      <div class="kpi-grid" id="kpis"></div>
      <div class="grid-2">
        <section class="panel">
          <h2>Language Mix</h2>
          <div id="languageChart"></div>
        </section>
        <section class="panel">
          <h2>Monthly Speech Trend</h2>
          <div id="timelineChart" class="timeline"></div>
        </section>
      </div>
      <div class="grid-2" style="margin-top:16px;">
        <section class="panel">
          <h2>Interaction Type Distribution</h2>
          <p class="panel-caption">Counts update with shared filters.</p>
          <div id="interactionChart" class="chart"></div>
        </section>
        <section class="panel">
          <h2>Field Domain Distribution</h2>
          <p class="panel-caption">Domain emphasis among the currently selected speeches.</p>
          <div id="domainChart" class="chart"></div>
        </section>
      </div>
      <section class="panel full-width-panel">
        <h2>Interaction by Domain</h2>
        <div id="heatmap" class="heatmap"></div>
      </section>
    </section>

    <section id="framework" class="view">
      <div class="panel">
        <h2>Situational Context Framework</h2>
        <p class="panel-caption">Framework entries from the workbook, connected to the filtered corpus counts.</p>
        <div class="framework-grid" id="frameworkGrid"></div>
        <div class="detail-box" id="frameworkDetail"></div>
      </div>
    </section>

    <section id="summary" class="view">
      <div class="summary-cards">
        <section class="panel">
          <h2>Corpus and Languages</h2>
          <div id="summaryCorpus"></div>
        </section>
        <section class="panel">
          <h2>Interaction Type</h2>
          <div id="summaryInteraction"></div>
        </section>
        <section class="panel">
          <h2>Field Domain</h2>
          <div id="summaryDomain"></div>
        </section>
      </div>
      <div class="full-width-panel">
        <section class="panel">
          <h2>Visual Mapping Counts</h2>
          <p class="panel-caption">Distribution table from the second workbook.</p>
          <div id="visualMapping" class="visual-map"></div>
        </section>
      </div>
    </section>

    <section id="analysis" class="view">
      <div class="analysis-toolbar">
        <span id="tableStatus"></span>
        <span>Click a column header to sort.</span>
      </div>
      <div class="table-wrap">
        <table id="analysisTable">
          <thead>
            <tr>
              <th data-key="no">No</th>
              <th data-key="date">Date</th>
              <th data-key="language">Language</th>
              <th data-key="event">Speech/Event</th>
              <th data-key="interactionType">Interaction Type</th>
              <th data-key="fieldDomain">Field Domain</th>
              <th data-key="wordCount">Words</th>
              <th data-key="openingEvidence">Opening Evidence</th>
            </tr>
          </thead>
          <tbody></tbody>
        </table>
      </div>
    </section>

    <section id="linguistic" class="view">
      <div class="research-head">
        <div>
          <h2>Linguistic Analysis</h2>
          <p>Explore how Prabowo positions self and collective voice through pronouns, appraisal, engagement, and first-person verb framing. Rates are normalized per 1,000 presidential words unless shown as percentages.</p>
        </div>
        <div class="sample-note" id="researchSampleNote"></div>
      </div>
      <div class="kpi-grid research-kpis" id="researchKpis"></div>
      <nav class="research-nav" aria-label="Linguistic analysis views">
        <button class="active" data-research-view="selfmention">Self-Mention</button>
        <button data-research-view="appraisal">Appraisal</button>
        <button data-research-view="engagement">Engagement</button>
        <button data-research-view="frames">Verb Frames</button>
        <button data-research-view="quality">Data Quality</button>
        <button data-research-view="sources">Source Tables</button>
      </nav>
      <section id="research-selfmention" class="research-panel active"></section>
      <section id="research-appraisal" class="research-panel"></section>
      <section id="research-engagement" class="research-panel"></section>
      <section id="research-frames" class="research-panel"></section>
      <section id="research-quality" class="research-panel"></section>
      <section id="research-sources" class="research-panel"></section>
    </section>
  </main>

  <script>
    const DATA = {data_json};

    const state = {{
      view: "overview",
      search: "",
      language: "All languages",
      year: "All years",
      interaction: "All interaction types",
      domain: "All field domains",
      sortKey: "date",
      sortDir: "asc",
      researchView: "selfmention",
      researchTable: "T0_per_speech_master",
      researchSearch: "",
      frameworkTerm: "Field",
      summarySort: {{
        corpus: {{ key: "order", dir: "asc" }},
        interaction: {{ key: "order", dir: "asc" }},
        domain: {{ key: "order", dir: "asc" }}
      }}
    }};

    const colors = ["#9a2b27", "#2f4858", "#876c34", "#4c5d4d", "#5a4a5c", "#3c3a36", "#7c5a3a", "#445a6a", "#6a5436"];

    const fmt = new Intl.NumberFormat("en-US");
    const pct = value => `${{Math.round((Number(value) || 0) * 1000) / 10}}%`;
    const shortDate = value => value ? new Date(value + "T00:00:00").toLocaleDateString("en-GB", {{ day: "2-digit", month: "short", year: "numeric" }}) : "";
    const esc = value => String(value ?? "").replace(/[&<>"']/g, char => ({{ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }}[char]));

    function uniq(values) {{
      return [...new Set(values.filter(Boolean))].sort((a, b) => String(a).localeCompare(String(b)));
    }}

    function countBy(rows, key) {{
      const map = new Map();
      rows.forEach(row => {{
        const value = row[key] || "Unspecified";
        map.set(value, (map.get(value) || 0) + 1);
      }});
      return [...map.entries()].map(([label, value]) => ({{ label, value }})).sort((a, b) => b.value - a.value || a.label.localeCompare(b.label));
    }}

    function sumBy(rows, key) {{
      return rows.reduce((total, row) => total + (Number(row[key]) || 0), 0);
    }}

    function average(rows, key) {{
      if (!rows.length) return 0;
      return Math.round(sumBy(rows, key) / rows.length);
    }}

    function fillSelect(id, label, values) {{
      const el = document.getElementById(id);
      const current = el.value || label;
      el.innerHTML = [`<option>${{label}}</option>`, ...values.map(value => `<option>${{esc(value)}}</option>`)].join("");
      el.value = values.includes(current) ? current : label;
    }}

    function initFilters() {{
      fillSelect("languageFilter", "All languages", uniq(DATA.analysis.map(d => d.language)));
      fillSelect("yearFilter", "All years", uniq(DATA.analysis.map(d => d.year)));
      fillSelect("interactionFilter", "All interaction types", uniq(DATA.analysis.map(d => d.interactionType)));
      fillSelect("domainFilter", "All field domains", uniq(DATA.analysis.map(d => d.fieldDomain)));

      document.getElementById("searchInput").addEventListener("input", e => {{ state.search = e.target.value.trim().toLowerCase(); render(); }});
      document.getElementById("languageFilter").addEventListener("change", e => {{ state.language = e.target.value; render(); }});
      document.getElementById("yearFilter").addEventListener("change", e => {{ state.year = e.target.value; render(); }});
      document.getElementById("interactionFilter").addEventListener("change", e => {{ state.interaction = e.target.value; render(); }});
      document.getElementById("domainFilter").addEventListener("change", e => {{ state.domain = e.target.value; render(); }});
      document.getElementById("clearFilters").addEventListener("click", () => {{
        state.search = "";
        state.language = "All languages";
        state.year = "All years";
        state.interaction = "All interaction types";
        state.domain = "All field domains";
        document.getElementById("searchInput").value = "";
        document.getElementById("languageFilter").value = state.language;
        document.getElementById("yearFilter").value = state.year;
        document.getElementById("interactionFilter").value = state.interaction;
        document.getElementById("domainFilter").value = state.domain;
        render();
      }});
    }}

    function filteredRows() {{
      return DATA.analysis.filter(row => {{
        const text = [row.event, row.interactionType, row.fieldDomain, row.field, row.tenor, row.mode, row.openingEvidence].join(" ").toLowerCase();
        return (!state.search || text.includes(state.search))
          && (state.language === "All languages" || row.language === state.language)
          && (state.year === "All years" || row.year === state.year)
          && (state.interaction === "All interaction types" || row.interactionType === state.interaction)
          && (state.domain === "All field domains" || row.fieldDomain === state.domain);
      }});
    }}

    function renderKpis(rows) {{
      const topInteraction = countBy(rows, "interactionType")[0];
      const topDomain = countBy(rows, "fieldDomain")[0];
      const dateValues = rows.map(row => row.date).filter(Boolean).sort();
      const kpis = [
        {{ label: "Speeches", value: fmt.format(rows.length), note: `${{fmt.format(DATA.analysis.length)}} in full corpus` }},
        {{ label: "Total Words", value: fmt.format(sumBy(rows, "wordCount")), note: `${{fmt.format(average(rows, "wordCount"))}} average words` }},
        {{ label: "Languages", value: fmt.format(uniq(rows.map(row => row.language)).length), note: countBy(rows, "language").map(d => `${{d.label}} ${{d.value}}`).join(" | ") || "No match" }},
        {{ label: "Top Interaction", value: topInteraction ? fmt.format(topInteraction.value) : "0", note: topInteraction ? topInteraction.label : "No match" }},
        {{ label: "Top Domain", value: topDomain ? fmt.format(topDomain.value) : "0", note: topDomain ? topDomain.label : "No match" }}
      ];
      document.getElementById("kpis").innerHTML = kpis.map(kpi => `
        <article class="kpi">
          <div class="kpi-label">${{esc(kpi.label)}}</div>
          <div class="kpi-value">${{esc(kpi.value)}}</div>
          <div class="kpi-note">${{esc(kpi.note)}}</div>
        </article>
      `).join("");
    }}

    function renderBars(id, rows, key, className = "") {{
      const data = countBy(rows, key);
      const max = Math.max(...data.map(d => d.value), 1);
      document.getElementById(id).innerHTML = data.length ? data.map(d => `
        <div class="bar-row" title="${{esc(d.label)}}: ${{d.value}}">
          <div class="bar-label">${{esc(d.label)}}</div>
          <div class="bar-track"><div class="bar-fill ${{className}}" style="width:${{Math.max(3, d.value / max * 100)}}%"></div></div>
          <div class="bar-value">${{fmt.format(d.value)}} <span>(${{pct(d.value / rows.length)}})</span></div>
        </div>
      `).join("") : `<div class="empty">No matching speeches</div>`;
    }}

    function renderLanguage(rows) {{
      const data = countBy(rows, "language");
      const total = rows.length || 1;
      let angle = 0;
      const stops = data.map((d, idx) => {{
        const start = angle;
        angle += (d.value / total) * 360;
        return `${{colors[idx % colors.length]}} ${{start}}deg ${{angle}}deg`;
      }}).join(", ");
      document.getElementById("languageChart").innerHTML = `
        <div class="donut-wrap">
          <div class="donut" style="background:${{stops ? `conic-gradient(${{stops}})` : "#edf2f7"}}"></div>
          <div class="legend">
            ${{data.map((d, idx) => `
              <div class="legend-item">
                <span class="swatch" style="background:${{colors[idx % colors.length]}}"></span>
                <span>${{esc(d.label)}}</span>
                <strong>${{fmt.format(d.value)}} (${{pct(d.value / total)}})</strong>
              </div>
            `).join("") || `<div class="empty">No matching speeches</div>`}}
          </div>
        </div>
      `;
    }}

    function renderTimeline(rows) {{
      const map = new Map();
      rows.forEach(row => {{
        if (!row.date) return;
        const key = row.date.slice(0, 7);
        map.set(key, (map.get(key) || 0) + 1);
      }});
      const data = [...map.entries()].sort().map(([label, value]) => ({{ label, value }}));
      const max = Math.max(...data.map(d => d.value), 1);
      if (!data.length) {{
        document.getElementById("timelineChart").innerHTML = `<div class="empty">No matching dates</div>`;
        return;
      }}
      const width = 720;
      const height = 248;
      const pad = {{ left: 42, right: 18, top: 24, bottom: 42 }};
      const plotWidth = width - pad.left - pad.right;
      const plotHeight = height - pad.top - pad.bottom;
      const bottomY = height - pad.bottom;
      const points = data.map((d, index) => {{
        const x = data.length === 1 ? pad.left + plotWidth / 2 : pad.left + (index / (data.length - 1)) * plotWidth;
        const y = bottomY - (d.value / max) * plotHeight;
        return {{ ...d, x, y }};
      }});
      const path = points.map((point, index) => `${{index ? "L" : "M"}} ${{point.x.toFixed(1)}} ${{point.y.toFixed(1)}}`).join(" ");
      const area = `${{path}} L ${{points[points.length - 1].x.toFixed(1)}} ${{bottomY}} L ${{points[0].x.toFixed(1)}} ${{bottomY}} Z`;
      const labelEvery = Math.max(1, Math.ceil(data.length / 7));
      const gridLines = [0, 0.5, 1].map(level => {{
        const y = bottomY - level * plotHeight;
        return `<line class="line-grid" x1="${{pad.left}}" x2="${{width - pad.right}}" y1="${{y}}" y2="${{y}}"></line>`;
      }}).join("");
      document.getElementById("timelineChart").innerHTML = `
        <svg class="line-chart" viewBox="0 0 ${{width}} ${{height}}" role="img" aria-label="Monthly speech trend line graph">
          ${{gridLines}}
          <line class="line-axis" x1="${{pad.left}}" x2="${{width - pad.right}}" y1="${{bottomY}}" y2="${{bottomY}}"></line>
          <line class="line-axis" x1="${{pad.left}}" x2="${{pad.left}}" y1="${{pad.top}}" y2="${{bottomY}}"></line>
          <path class="line-area" d="${{area}}"></path>
          <path class="line-path" d="${{path}}"></path>
          ${{points.map(point => `<circle class="line-point" cx="${{point.x.toFixed(1)}}" cy="${{point.y.toFixed(1)}}" r="5"><title>${{esc(point.label)}}: ${{point.value}}</title></circle>`).join("")}}
          ${{points.map((point, index) => index % labelEvery === 0 || index === points.length - 1 ? `<text class="line-label" x="${{point.x.toFixed(1)}}" y="${{height - 12}}" text-anchor="middle">${{esc(point.label.slice(5) + "/" + point.label.slice(2, 4))}}</text>` : "").join("")}}
          ${{points.map(point => `<text class="line-value" x="${{point.x.toFixed(1)}}" y="${{Math.max(14, point.y - 12).toFixed(1)}}" text-anchor="middle">${{fmt.format(point.value)}}</text>`).join("")}}
        </svg>
      `;
    }}

    function renderHeatmap(rows) {{
      const interactions = countBy(rows, "interactionType").map(d => d.label);
      const domains = countBy(rows, "fieldDomain").map(d => d.label);
      const max = Math.max(...interactions.flatMap(i => domains.map(d => rows.filter(row => row.interactionType === i && row.fieldDomain === d).length)), 1);
      const titleCase = text => String(text || "").replace(/\\w\\S*/g, word => word.charAt(0).toUpperCase() + word.slice(1).toLowerCase());
      const head = `<tr><th>Interaction</th>${{domains.map(d => `<th>${{esc(titleCase(d))}}</th>`).join("")}}</tr>`;
      const body = interactions.map(interaction => `
        <tr>
          <td><strong>${{esc(titleCase(interaction))}}</strong></td>
          ${{domains.map(domain => {{
            const value = rows.filter(row => row.interactionType === interaction && row.fieldDomain === domain).length;
            const alpha = value ? 0.12 + (value / max) * 0.62 : 0;
            return `<td style="background:rgba(140,42,39,${{alpha}}); color:${{alpha > 0.55 ? "#f4f1ea" : "inherit"}}; text-align:center; font-variant-numeric:tabular-nums;">${{value || ""}}</td>`;
          }}).join("")}}
        </tr>
      `).join("");
      document.getElementById("heatmap").innerHTML = rows.length ? `<table class="heatmap-table">${{head}}${{body}}</table>` : `<div class="empty">No matching speeches</div>`;
    }}

    function renderFramework(rows) {{
      const focusCounts = {{
        Field: countBy(rows, "fieldDomain").slice(0, 4).map(d => `${{d.label}} (${{d.value}})`).join("; "),
        Tenor: countBy(rows, "interactionType").slice(0, 4).map(d => `${{d.label}} (${{d.value}})`).join("; "),
        Mode: countBy(rows, "language").map(d => `${{d.label}} (${{d.value}})`).join("; ")
      }};
      document.getElementById("frameworkGrid").innerHTML = DATA.framework.map(item => `
        <article class="framework-item ${{item.term === state.frameworkTerm ? "active" : ""}}" data-term="${{esc(item.term)}}">
          <h3>${{esc(item.term)}}</h3>
          <p>${{esc(item.meaning || "")}}</p>
          <p style="margin-top:10px;"><strong>${{esc(item.components || "")}}</strong></p>
        </article>
      `).join("");
      document.querySelectorAll(".framework-item").forEach(item => item.addEventListener("click", () => {{
        state.frameworkTerm = item.dataset.term;
        renderFramework(DATA.analysis);
      }}));
      const selected = DATA.framework.find(item => item.term === state.frameworkTerm) || DATA.framework[0];
      document.getElementById("frameworkDetail").innerHTML = selected ? `
        <h3>${{esc(selected.term)}}</h3>
        <p>${{esc(selected.meaning || "")}}</p>
        <p><strong>Workbook components:</strong> ${{esc(selected.components || "")}}</p>
        <p><strong>Full corpus signal:</strong> ${{esc(focusCounts[selected.term] || `${{rows.length}} matching speeches`)}}</p>
      ` : "";
    }}

    function renderSummary() {{
      const sortSummaryRows = (group, rows) => {{
        const sort = state.summarySort[group] || {{ key: "order", dir: "asc" }};
        return [...rows].sort((a, b) => {{
          let av = a[sort.key];
          let bv = b[sort.key];
          if (sort.key === "count" || sort.key === "share" || sort.key === "order") {{
            av = sort.key === "share" ? Number(a.share) || (String(a.share).includes("100") ? 1 : 0) : Number(av) || 0;
            bv = sort.key === "share" ? Number(b.share) || (String(b.share).includes("100") ? 1 : 0) : Number(bv) || 0;
          }} else {{
            av = String(av || "").toLowerCase();
            bv = String(bv || "").toLowerCase();
          }}
          if (av < bv) return sort.dir === "asc" ? -1 : 1;
          if (av > bv) return sort.dir === "asc" ? 1 : -1;
          return 0;
        }});
      }};

      const sortMark = (group, key) => {{
        const sort = state.summarySort[group];
        return sort && sort.key === key ? (sort.dir === "asc" ? " ↑" : " ↓") : "";
      }};

      const summaryTable = (group, rows) => `
        <table class="summary-table">
          <thead>
            <tr>
              <th class="${{state.summarySort[group].key === "category" ? "active-sort" : ""}}" data-summary-group="${{group}}" data-summary-key="category">Category${{sortMark(group, "category")}}</th>
              <th class="${{state.summarySort[group].key === "count" ? "active-sort" : ""}}" data-summary-group="${{group}}" data-summary-key="count">Count${{sortMark(group, "count")}}</th>
              <th class="${{state.summarySort[group].key === "share" ? "active-sort" : ""}}" data-summary-group="${{group}}" data-summary-key="share">Share${{sortMark(group, "share")}}</th>
            </tr>
          </thead>
          <tbody>
            ${{sortSummaryRows(group, rows).map(row => `
              <tr>
                <td>${{esc(row.category || "")}}</td>
                <td>${{fmt.format(Number(row.count) || 0)}}</td>
                <td>${{typeof row.share === "number" ? pct(row.share) : esc(row.share || "")}}</td>
              </tr>
            `).join("")}}
          </tbody>
        </table>
      `;

      const summaryWithOrder = DATA.summary.map((row, order) => ({{ ...row, order }}));
      document.getElementById("summaryCorpus").innerHTML = summaryTable("corpus", summaryWithOrder.filter(row => row.metric === "Total speeches" || row.metric === "Language group"));
      document.getElementById("summaryInteraction").innerHTML = summaryTable("interaction", summaryWithOrder.filter(row => row.metric === "Interaction type"));
      document.getElementById("summaryDomain").innerHTML = summaryTable("domain", summaryWithOrder.filter(row => row.metric === "Field domain"));

      const interaction = DATA.visual.interactionTypes;
      const domains = DATA.visual.fieldDomains;
      document.getElementById("visualMapping").innerHTML = `
        <h3>Interaction Types</h3>
        ${{interaction.map(item => `
          <div class="bar-row">
            <div class="bar-label">${{esc(item.category)}}</div>
            <div class="bar-track"><div class="bar-fill coral" style="width:${{Math.max(3, Number(item.share || 0) * 100)}}%"></div></div>
            <div class="bar-value">${{fmt.format(item.count)}} (${{pct(item.share)}})</div>
          </div>
        `).join("")}}
        <h3 style="margin-top:22px;">Field Domains</h3>
        ${{domains.map(item => `
          <div class="bar-row">
            <div class="bar-label">${{esc(item.category)}}</div>
            <div class="bar-track"><div class="bar-fill indigo" style="width:${{Math.max(3, Number(item.share || 0) * 100)}}%"></div></div>
            <div class="bar-value">${{fmt.format(item.count)}} (${{pct(item.share)}})</div>
          </div>
        `).join("")}}
      `;
    }}

    function sortedRows(rows) {{
      return [...rows].sort((a, b) => {{
        let av = a[state.sortKey];
        let bv = b[state.sortKey];
        if (state.sortKey === "wordCount") {{
          av = Number(av) || 0;
          bv = Number(bv) || 0;
        }} else {{
          av = String(av || "");
          bv = String(bv || "");
        }}
        if (av < bv) return state.sortDir === "asc" ? -1 : 1;
        if (av > bv) return state.sortDir === "asc" ? 1 : -1;
        return 0;
      }});
    }}

    function renderTable(rows) {{
      const tbody = document.querySelector("#analysisTable tbody");
      const sorted = sortedRows(rows);
      document.getElementById("tableStatus").textContent = `${{fmt.format(sorted.length)}} matching speeches`;
      tbody.innerHTML = sorted.map(row => `
        <tr>
          <td>${{esc(row.no)}}</td>
          <td>${{esc(shortDate(row.date))}}</td>
          <td>${{esc(row.language)}}</td>
          <td><strong>${{esc(row.event)}}</strong></td>
          <td>${{esc(row.interactionType)}}</td>
          <td>${{esc(row.fieldDomain)}}</td>
          <td style="text-align:right; font-variant-numeric:tabular-nums;">${{fmt.format(row.wordCount)}}</td>
          <td class="evidence">${{esc(row.openingEvidence)}}</td>
        </tr>
      `).join("");
    }}

    function researchRows(key) {{
      return DATA.research?.tables?.[key]?.rows || [];
    }}

    function formatMetric(value, digits = 2) {{
      const number = Number(value);
      if (!Number.isFinite(number)) return esc(value);
      return number.toLocaleString("en-US", {{ maximumFractionDigits: digits }});
    }}

    function renderResearchKpis() {{
      const full = researchRows("T1b_selfmention_EXCLUSIONS_APPLIED");
      const fullTotals = full.filter(row => row.form === "TOTAL");
      const sample = researchRows("T2_selfmention_sample");
      const sampleTotals = sample.filter(row => row.form === "TOTAL");
      const hygiene = researchRows("T8_data_hygiene_log").filter(row => Number(row.words_removed) > 0);
      const corpusTexts = fullTotals.reduce((sum, row) => sum + (Number(row.texts) || 0), 0);
      const corpusWords = fullTotals.reduce((sum, row) => sum + (Number(row.running_words) || 0), 0);
      const sampleWords = sampleTotals.reduce((sum, row) => sum + (Number(row.running_words) || 0), 0);
      const removedWords = hygiene.reduce((sum, row) => sum + (Number(row.words_removed) || 0), 0);
      const kpis = [
        {{ label: "Clean Analysis Corpus", value: fmt.format(corpusTexts), note: `${{fmt.format(corpusWords)}} words after exclusions` }},
        {{ label: "Purposive Sample", value: fmt.format((DATA.research.samples.domestic || []).length + (DATA.research.samples.international || []).length), note: `${{fmt.format(sampleWords)}} words across two samples` }},
        {{ label: "Analysis Tables", value: fmt.format(Object.keys(DATA.research.tables).length), note: "All available in Source Tables" }},
        {{ label: "Cleaned Transcripts", value: fmt.format(hygiene.length), note: `${{fmt.format(removedWords)}} non-Prabowo words removed` }},
        {{ label: "Sample Exclusions", value: fmt.format(Object.keys(DATA.research.samples.excluded || {{}}).length), note: "Duplicate or translated speeches" }}
      ];
      document.getElementById("researchKpis").innerHTML = kpis.map(kpi => `
        <article class="kpi">
          <div class="kpi-label">${{esc(kpi.label)}}</div>
          <div class="kpi-value">${{esc(kpi.value)}}</div>
          <div class="kpi-note">${{esc(kpi.note)}}</div>
        </article>
      `).join("");
      document.getElementById("researchSampleNote").innerHTML = `Domestic sample: <strong>${{(DATA.research.samples.domestic || []).length}} speeches</strong><br>International sample: <strong>${{(DATA.research.samples.international || []).length}} speeches</strong>`;
    }}

    function rateRows(rows, valueKey = "per_1000w", colorClass = "") {{
      const max = Math.max(...rows.map(row => Number(row[valueKey]) || 0), 1);
      return `<div class="rate-chart">${{rows.map(row => `
        <div class="rate-row">
          <div class="rate-label"><strong>${{esc(row.form || row.frame || row.pronoun_node)}}</strong><span>${{row.function ? `${{esc(row.function)}} · ` : ""}}${{fmt.format(Number(row.raw) || Number(row.nodes) || 0)}} raw</span></div>
          <div class="bar-track"><div class="bar-fill ${{colorClass}}" style="width:${{Math.max(2, (Number(row[valueKey]) || 0) / max * 100)}}%"></div></div>
          <div class="rate-value">${{formatMetric(row[valueKey])}} / 1k</div>
        </div>
      `).join("")}}</div>`;
    }}

    function renderSelfMention() {{
      const rows = researchRows("T2_selfmention_sample").filter(row => row.form !== "TOTAL");
      const domestic = rows.filter(row => String(row.sample).startsWith("Domestic"));
      const international = rows.filter(row => String(row.sample).startsWith("International"));
      const sumRate = (items, forms) => items.filter(row => forms.includes(row.form)).reduce((sum, row) => sum + (Number(row.per_1000w) || 0), 0);
      const domesticCollective = sumRate(domestic, ["kami", "kita"]);
      const domesticSingular = sumRate(domestic, ["saya"]);
      const internationalCollective = sumRate(international, ["we", "us", "our"]);
      const internationalSingular = sumRate(international, ["I", "me", "my"]);
      document.getElementById("research-selfmention").innerHTML = `
        <div class="analysis-grid">
          <section class="panel">
            <h2>Domestic Sample</h2>
            <p class="panel-caption">Bahasa Indonesia, n=${{DATA.research.samples.domestic.length}}. Pronoun frequency per 1,000 words.</p>
            ${{rateRows(domestic, "per_1000w", "coral")}}
          </section>
          <section class="panel">
            <h2>International Sample</h2>
            <p class="panel-caption">English, n=${{DATA.research.samples.international.length}}. Pronoun frequency per 1,000 words.</p>
            ${{rateRows(international, "per_1000w", "slate")}}
          </section>
          <section class="panel wide">
            <h2>Collective vs. Singular Voice</h2>
            <div class="grid-2">
              <div class="finding"><strong>${{formatMetric(domesticCollective)}}</strong> collective mentions per 1,000 words (<em>kami + kita</em>), compared with ${{formatMetric(domesticSingular)}} for <em>saya</em>.</div>
              <div class="finding"><strong>${{formatMetric(internationalCollective)}}</strong> collective mentions per 1,000 words (<em>we + us + our</em>), compared with ${{formatMetric(internationalSingular)}} for <em>I + me + my</em>.</div>
            </div>
          </section>
        </div>`;
    }}

    function renderAppraisal() {{
      const rows = researchRows("T5b_attitude_group_summary");
      const palette = {{ AFF_pct: "var(--accent)", JUD_pct: "var(--slate)", APP_pct: "var(--ochre)" }};
      document.getElementById("research-appraisal").innerHTML = `
        <div class="analysis-grid">
          <section class="panel wide">
            <h2>Attitude Group Profile</h2>
            <p class="panel-caption">Share of coded attitude surrounding each pronoun node.</p>
            <div class="analysis-legend"><span><i class="swatch" style="background:${{palette.AFF_pct}}"></i>Affect</span><span><i class="swatch" style="background:${{palette.JUD_pct}}"></i>Judgement</span><span><i class="swatch" style="background:${{palette.APP_pct}}"></i>Appreciation</span></div>
            ${{rows.map(row => `
              <div class="attitude-row">
                <div class="attitude-row-head"><strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}}</strong><span>${{formatMetric(row.attitude_per_1000w)}} / 1k · ${{fmt.format(row.nodes)}} nodes</span></div>
                <div class="stacked-bar" title="Affect ${{row.AFF_pct}}%, Judgement ${{row.JUD_pct}}%, Appreciation ${{row.APP_pct}}%">
                  <i style="width:${{row.AFF_pct}}%;background:${{palette.AFF_pct}}"></i><i style="width:${{row.JUD_pct}}%;background:${{palette.JUD_pct}}"></i><i style="width:${{row.APP_pct}}%;background:${{palette.APP_pct}}"></i>
                </div>
              </div>`).join("")}}
          </section>
          <section class="panel">
            <h2>Dominant Attitude</h2>
            ${{rows.map(row => {{
              const groups = [["Affect", Number(row.AFF_pct)], ["Judgement", Number(row.JUD_pct)], ["Appreciation", Number(row.APP_pct)]].sort((a, b) => b[1] - a[1]);
              return `<div class="finding"><strong>${{formatMetric(groups[0][1], 1)}}%</strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}} is led by ${{groups[0][0].toLowerCase()}}.</div>`;
            }}).join("")}}
          </section>
          <section class="panel">
            <h2>Highest Attitude Density</h2>
            ${{[...rows].sort((a, b) => Number(b.attitude_per_1000w) - Number(a.attitude_per_1000w)).map(row => `<div class="finding"><strong>${{formatMetric(row.attitude_per_1000w)}}</strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}} per 1,000 words.</div>`).join("")}}
          </section>
        </div>`;
    }}

    function renderEngagement() {{
      const rows = researchRows("T6_engagement_hedge_booster");
      const max = Math.max(...rows.flatMap(row => [Number(row.hedges_per1k), Number(row.boosters_per1k)]), 1);
      document.getElementById("research-engagement").innerHTML = `
        <div class="analysis-grid">
          <section class="panel wide">
            <h2>Hedges and Boosters</h2>
            <p class="panel-caption">Normalized frequencies around each pronoun node.</p>
            <div class="analysis-legend"><span><i class="swatch" style="background:var(--ochre)"></i>Hedges</span><span><i class="swatch" style="background:var(--slate)"></i>Boosters</span></div>
            <div class="rate-chart">${{rows.map(row => `
              <div class="rate-row">
                <div class="rate-label"><strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}}</strong><span>ratio ${{formatMetric(row.hedge_booster_ratio)}}</span></div>
                <div class="grouped-track">
                  <div class="bar-track" title="Hedges: ${{row.hedges_per1k}} per 1,000"><div class="bar-fill ochre" style="width:${{Math.max(2, Number(row.hedges_per1k) / max * 100)}}%"></div></div>
                  <div class="bar-track" title="Boosters: ${{row.boosters_per1k}} per 1,000"><div class="bar-fill slate" style="width:${{Math.max(2, Number(row.boosters_per1k) / max * 100)}}%"></div></div>
                </div>
                <div class="rate-value">${{formatMetric(row.hedges_per1k)}} / ${{formatMetric(row.boosters_per1k)}}</div>
              </div>`).join("")}}</div>
          </section>
          <section class="panel wide">
            <h2>Engagement Balance</h2>
            <div class="grid-2">
              ${{rows.map(row => `<div class="finding"><strong>${{Number(row.hedge_booster_ratio) > 1 ? "H" : "B"}}</strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}} is ${{Number(row.hedge_booster_ratio) > 1 ? "hedge-led" : "booster-led"}} (ratio ${{formatMetric(row.hedge_booster_ratio)}}).</div>`).join("")}}
            </div>
          </section>
        </div>`;
    }}

    function renderVerbFrames() {{
      const rows = researchRows("T7_firstperson_verb_frames");
      const domestic = rows.filter(row => String(row.sample).startsWith("Domestic"));
      const international = rows.filter(row => String(row.sample).startsWith("International"));
      const byFunction = items => {{
        const groups = new Map();
        items.forEach(row => {{
          const current = groups.get(row.function) || {{ frame: row.function, raw: 0, per_1000w: 0 }};
          current.raw += Number(row.raw) || 0;
          current.per_1000w += Number(row.per_1000w) || 0;
          groups.set(row.function, current);
        }});
        return [...groups.values()].sort((a, b) => b.per_1000w - a.per_1000w);
      }};
      const functionPanel = (title, items, color) => `
        <section class="panel"><h2>${{title}}</h2><p class="panel-caption">Combined rate of frames assigned to each discourse function.</p>${{rateRows(byFunction(items), "per_1000w", color)}}</section>`;
      const framePanel = (title, items, color) => `
        <section class="panel"><h2>${{title}}</h2><p class="panel-caption">Individual first-person verb frames, ordered by normalized frequency.</p>${{rateRows([...items].sort((a, b) => Number(b.per_1000w) - Number(a.per_1000w)), "per_1000w", color)}}</section>`;
      document.getElementById("research-frames").innerHTML = `<div class="analysis-grid">
        ${{functionPanel("Domestic Functional Profile", domestic, "coral")}}
        ${{functionPanel("International Functional Profile", international, "slate")}}
        ${{framePanel("Domestic Verb Frames", domestic, "coral")}}
        ${{framePanel("International Verb Frames", international, "slate")}}
      </div>`;
    }}

    function renderQuality() {{
      const rows = researchRows("T8_data_hygiene_log");
      const cleaned = rows.filter(row => Number(row.words_removed) > 0);
      const removed = cleaned.reduce((sum, row) => sum + Number(row.words_removed), 0);
      const raw = cleaned.reduce((sum, row) => sum + Number(row.raw_words), 0);
      const exclusions = Object.entries(DATA.research.samples.excluded || {{}});
      const top = [...cleaned].sort((a, b) => Number(b.pct_removed) - Number(a.pct_removed)).slice(0, 10);
      document.getElementById("research-quality").innerHTML = `
        <div class="analysis-grid">
          <section class="panel">
            <h2>Speaker-Turn Cleaning</h2>
            <div class="finding"><strong>${{fmt.format(cleaned.length)}}</strong>multi-speaker transcripts cleaned.</div>
            <div class="finding"><strong>${{fmt.format(removed)}}</strong>non-Prabowo words removed.</div>
            <div class="finding"><strong>${{pct(removed / Math.max(raw, 1))}}</strong>of words removed within affected transcripts.</div>
          </section>
          <section class="panel">
            <h2>Purposive-Sample Exclusions</h2>
            ${{exclusions.map(([id, reason]) => `<div class="finding"><strong>${{esc(id)}}</strong>${{esc(reason)}}</div>`).join("")}}
          </section>
          <section class="panel wide">
            <h2>Largest Speaker-Turn Reductions</h2>
            <div class="table-wrap" style="max-height:430px"><table class="summary-table"><thead><tr><th>Speech</th><th>Language</th><th>Presidential Words</th><th>Words Removed</th><th>Removed</th></tr></thead><tbody>${{top.map(row => `<tr><td>${{esc(row.id)}}</td><td>${{esc(row.lang)}}</td><td>${{fmt.format(row.prabowo_words)}}</td><td>${{fmt.format(row.words_removed)}}</td><td>${{formatMetric(row.pct_removed, 1)}}%</td></tr>`).join("")}}</tbody></table></div>
          </section>
        </div>`;
    }}

    function prettyHeader(value) {{
      return String(value).replaceAll("_", " ").replace(/\b\w/g, char => char.toUpperCase());
    }}

    function sourceCell(value) {{
      if (value === true) return "Yes";
      if (value === false) return "No";
      if (value === "" || value == null) return "—";
      return value;
    }}

    function renderResearchSources() {{
      const tables = DATA.research.tables;
      const keys = Object.keys(tables);
      if (!tables[state.researchTable]) state.researchTable = keys[0];
      const table = tables[state.researchTable];
      const query = state.researchSearch.toLowerCase();
      const rows = table.rows.filter(row => !query || Object.values(row).some(value => String(value).toLowerCase().includes(query)));
      const headers = table.rows.length ? Object.keys(table.rows[0]) : [];
      const isMaster = state.researchTable === "T0_per_speech_master";
      const masterGroups = isMaster ? `<tr class="master-groups"><th colspan="3">Identity</th><th colspan="2">Discourse Context</th><th colspan="3">Corpus Words</th><th colspan="2">Sample Membership</th><th colspan="9">Raw Self-Mentions</th><th colspan="5">Normalized Rate per 1,000</th><th colspan="1">Speech</th></tr>` : "";
      const masterCols = isMaster ? `<colgroup>${{headers.map(() => "<col>").join("")}}</colgroup>` : "";
      document.getElementById("research-sources").innerHTML = `
        <section class="panel">
          <h2>Source Table Explorer</h2>
          <div class="source-toolbar">
            <select id="researchTableSelect" aria-label="Analysis source table">${{keys.map(key => `<option value="${{esc(key)}}" ${{key === state.researchTable ? "selected" : ""}}>${{esc(tables[key].title)}}</option>`).join("")}}</select>
            <input id="researchTableSearch" type="search" value="${{esc(state.researchSearch)}}" placeholder="Search this table" />
            <button class="clear-btn" id="researchTableClear">Clear</button>
          </div>
          <p class="source-meta">${{esc(table.source)}} · ${{fmt.format(rows.length)}} of ${{fmt.format(table.rows.length)}} rows</p>
          <div class="table-wrap" style="max-height:620px"><table class="source-table ${{isMaster ? "master-table" : ""}}">${{masterCols}}<thead>${{masterGroups}}<tr>${{headers.map(header => `<th>${{esc(prettyHeader(header))}}</th>`).join("")}}</tr></thead><tbody>${{rows.map(row => `<tr>${{headers.map(header => `<td class="${{typeof row[header] === "number" ? "numeric" : ""}}">${{esc(sourceCell(row[header]))}}</td>`).join("")}}</tr>`).join("")}}</tbody></table></div>
        </section>`;
      document.getElementById("researchTableSelect").addEventListener("change", event => {{ state.researchTable = event.target.value; state.researchSearch = ""; renderResearchSources(); }});
      document.getElementById("researchTableSearch").addEventListener("input", event => {{ state.researchSearch = event.target.value; renderResearchSources(); document.getElementById("researchTableSearch").focus(); }});
      document.getElementById("researchTableClear").addEventListener("click", () => {{ state.researchSearch = ""; renderResearchSources(); }});
    }}

    function renderResearch() {{
      renderResearchKpis();
      renderSelfMention();
      renderAppraisal();
      renderEngagement();
      renderVerbFrames();
      renderQuality();
      renderResearchSources();
    }}

    function bindResearchNav() {{
      document.querySelectorAll("[data-research-view]").forEach(button => button.addEventListener("click", () => {{
        state.researchView = button.dataset.researchView;
        document.querySelectorAll("[data-research-view]").forEach(item => item.classList.toggle("active", item === button));
        document.querySelectorAll(".research-panel").forEach(panel => panel.classList.toggle("active", panel.id === `research-${{state.researchView}}`));
        prepareScrollAnimations();
      }}));
    }}

    function bindTabs() {{
      document.querySelectorAll(".tab-btn").forEach(btn => btn.addEventListener("click", () => {{
        state.view = btn.dataset.view;
        document.querySelectorAll(".tab-btn").forEach(item => item.classList.toggle("active", item === btn));
        document.querySelectorAll(".view").forEach(view => view.classList.toggle("active", view.id === state.view));
        updateFilterVisibility();
        prepareScrollAnimations();
      }}));
    }}

    function updateFilterVisibility() {{
      const hideFilters = state.view === "framework" || state.view === "summary" || state.view === "linguistic";
      document.querySelector(".filters").classList.toggle("is-hidden", hideFilters);
    }}

    function bindTableSort() {{
      document.querySelectorAll("#analysisTable th").forEach(th => th.addEventListener("click", () => {{
        const key = th.dataset.key;
        if (state.sortKey === key) {{
          state.sortDir = state.sortDir === "asc" ? "desc" : "asc";
        }} else {{
          state.sortKey = key;
          state.sortDir = key === "wordCount" ? "desc" : "asc";
        }}
        renderTable(filteredRows());
      }}));
    }}

    function bindSummarySort() {{
      document.getElementById("summary").addEventListener("click", event => {{
        const th = event.target.closest("th[data-summary-key]");
        if (!th) return;
        const group = th.dataset.summaryGroup;
        const key = th.dataset.summaryKey;
        const current = state.summarySort[group];
        state.summarySort[group] = {{
          key,
          dir: current.key === key && current.dir === "asc" ? "desc" : "asc"
        }};
        renderSummary();
        prepareScrollAnimations();
      }});
    }}

    let revealObserver = null;

    function prepareScrollAnimations() {{
      if (window.matchMedia("(prefers-reduced-motion: reduce)").matches) return;
      document.body.classList.add("motion-ready");
      if (!revealObserver) {{
        revealObserver = new IntersectionObserver(entries => {{
          entries.forEach(entry => {{
            if (entry.isIntersecting) {{
              entry.target.classList.add("in-view");
              revealObserver.unobserve(entry.target);
            }}
          }});
        }}, {{ threshold: 0.16, rootMargin: "0px 0px -8% 0px" }});
      }}
      const activeView = document.querySelector(".view.active");
      const targets = [
        ...document.querySelectorAll(".kpi"),
        ...(activeView ? [...activeView.querySelectorAll(".panel, .framework-item")] : [])
      ];
      targets.forEach((el, index) => {{
        if (el.dataset.revealBound === "true") return;
        el.classList.add("reveal");
        el.style.setProperty("--reveal-delay", `${{Math.min(index * 45, 260)}}ms`);
        el.dataset.revealBound = "true";
        revealObserver.observe(el);
      }});
    }}

    function bindScrollMicroMotion() {{
      if (window.matchMedia("(prefers-reduced-motion: reduce)").matches) return;
      const masthead = document.querySelector(".masthead");
      let ticking = false;
      const update = () => {{
        const offset = Math.min(window.scrollY * 0.08, 18);
        masthead.style.transform = `translateY(${{offset}}px)`;
        ticking = false;
      }};
      window.addEventListener("scroll", () => {{
        if (!ticking) {{
          window.requestAnimationFrame(update);
          ticking = true;
        }}
      }}, {{ passive: true }});
    }}

    function render() {{
      const rows = filteredRows();
      renderKpis(rows);
      renderBars("interactionChart", rows, "interactionType", "coral");
      renderBars("domainChart", rows, "fieldDomain", "indigo");
      renderLanguage(rows);
      renderTimeline(rows);
      renderHeatmap(rows);
      renderFramework(DATA.analysis);
      renderSummary();
      renderTable(rows);
      renderResearch();
      updateFilterVisibility();
      prepareScrollAnimations();
    }}

    initFilters();
    bindTabs();
    bindTableSort();
    bindSummarySort();
    bindResearchNav();
    render();
  </script>
</body>
</html>
"""


def build_mobile_html(data: dict[str, Any]) -> str:
    data_json = json.dumps(data, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Prabowo Speeches Mobile Dashboard</title>
  <style>
    :root {{
      --paper:    #f4f1ea;
      --paper-2:  #ebe5d8;
      --panel:    #fbf9f3;
      --ink:      #1b1a16;
      --ink-soft: #3b3731;
      --muted:    #6b6458;
      --faint:    #9a9286;
      --line:     #dcd5c6;
      --line-2:   #c8c0ad;
      --rule:     #26231d;
      --accent:   #9a2b27;
      --slate:    #2f4858;
      --ochre:    #876c34;
      --serif: "Iowan Old Style", "Palatino Linotype", Palatino, Charter, Georgia, "Times New Roman", serif;
      --sans: -apple-system, BlinkMacSystemFont, "Helvetica Neue", "Segoe UI", Roboto, Arial, sans-serif;
      --mono: "SF Mono", ui-monospace, "DejaVu Sans Mono", "Roboto Mono", monospace;
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      min-width: 320px;
      font-family: var(--sans);
      background: var(--paper);
      color: var(--ink);
      padding-bottom: 84px;
      -webkit-font-smoothing: antialiased;
    }}

    header {{
      padding: 26px 18px 22px;
      color: var(--ink);
      background: var(--paper);
      border-top: 3px solid var(--accent);
      border-bottom: 1px solid var(--rule);
    }}

    .kicker {{
      margin: 0 0 12px;
      color: var(--accent);
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.16em;
    }}

    h1 {{
      margin: 0;
      font-family: var(--serif);
      font-size: clamp(30px, 8.6vw, 44px);
      font-weight: 600;
      line-height: 1.04;
      letter-spacing: -0.015em;
      overflow-wrap: break-word;
      color: var(--ink);
    }}

    .subtitle {{
      margin: 14px 0 0;
      font-family: var(--serif);
      font-size: 16px;
      line-height: 1.5;
      color: var(--muted);
    }}

    .mobile-nav {{
      position: fixed;
      left: 0;
      right: 0;
      bottom: 0;
      z-index: 20;
      display: grid;
      grid-template-columns: repeat(5, 1fr);
      gap: 0;
      padding: 0;
      border-top: 1px solid var(--rule);
      background: var(--paper);
    }}

    .mobile-nav button {{
      min-height: 56px;
      border: 0;
      border-right: 1px solid var(--line);
      background: transparent;
      color: var(--muted);
      font: inherit;
      font-size: 12px;
      font-weight: 600;
      letter-spacing: 0.02em;
    }}

    .mobile-nav button:last-child {{ border-right: 0; }}

    .mobile-nav button.active {{
      color: var(--ink);
      box-shadow: inset 0 -3px 0 var(--accent);
    }}

    main {{ padding: 18px 14px 0; }}

    .view {{ display: none; }}
    .view.active {{ display: block; }}

    .mobile-filters {{ display: grid; gap: 8px; margin-bottom: 16px; }}

    input, select {{
      width: 100%;
      min-height: 46px;
      border: 1px solid var(--line-2);
      border-radius: 2px;
      background: var(--panel);
      color: var(--ink);
      padding: 0 12px;
      font: inherit;
      font-size: 15px;
    }}

    input:focus, select:focus {{
      outline: none;
      border-color: var(--accent);
      box-shadow: inset 0 0 0 1px var(--accent);
    }}

    .reset {{
      min-height: 46px;
      border: 1px solid var(--rule);
      border-radius: 2px;
      background: transparent;
      color: var(--ink);
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      font-size: 13px;
    }}

    .card {{
      margin-bottom: 16px;
      padding: 18px;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: var(--panel);
    }}

    .card h2 {{
      margin: 0 0 14px;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
      font-family: var(--serif);
      font-size: 21px;
      font-weight: 600;
      line-height: 1.2;
      color: var(--ink);
    }}

    .card h3 {{
      margin: 18px 0 8px;
      font-family: var(--serif);
      font-size: 15px;
      font-weight: 600;
      color: var(--ink);
    }}

    .kpis {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 10px;
      margin-bottom: 16px;
    }}

    .kpi {{
      padding: 14px;
      min-height: 96px;
      border: 1px solid var(--line);
      border-top: 2px solid var(--rule);
      border-radius: 3px;
      background: var(--panel);
    }}

    .kpi span {{
      display: block;
      color: var(--muted);
      font-size: 10.5px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.1em;
    }}

    .kpi strong {{
      display: block;
      margin-top: 10px;
      font-family: var(--serif);
      font-weight: 600;
      font-size: 30px;
      line-height: 1;
      font-variant-numeric: tabular-nums;
    }}

    .kpi small {{
      display: block;
      margin-top: 8px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
    }}

    .bar-row {{
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 8px;
      margin: 14px 0;
      align-items: center;
    }}

    .bar-label {{ font-size: 13px; line-height: 1.35; color: var(--ink-soft); }}

    .bar-value {{
      color: var(--muted);
      font-family: var(--mono);
      font-size: 12px;
      font-variant-numeric: tabular-nums;
      white-space: nowrap;
    }}

    .bar-track {{
      grid-column: 1 / -1;
      height: 10px;
      border-radius: 2px;
      background: var(--paper-2);
      overflow: hidden;
    }}

    .bar-fill {{
      height: 100%;
      border-radius: inherit;
      background: var(--accent);
    }}

    .bar-fill.indigo {{ background: var(--slate); }}

    .donut-wrap {{ display: grid; justify-items: center; gap: 16px; }}

    .donut {{
      width: min(62vw, 220px);
      aspect-ratio: 1;
      border-radius: 50%;
      position: relative;
    }}

    .donut::after {{
      content: "";
      position: absolute;
      inset: 28%;
      border-radius: 50%;
      background: var(--panel);
    }}

    .legend {{ width: 100%; display: grid; gap: 0; }}

    .legend-row {{
      display: grid;
      grid-template-columns: 10px 1fr auto;
      gap: 10px;
      align-items: center;
      font-size: 13px;
      padding: 9px 0;
      border-bottom: 1px solid var(--line);
    }}

    .legend-row:last-child {{ border-bottom: 0; }}

    .legend-row strong {{ font-family: var(--mono); font-weight: 600; color: var(--ink); }}

    .swatch {{ width: 10px; height: 10px; border-radius: 0; }}

    .line-chart {{ width: 100%; height: auto; display: block; }}

    .line-grid {{ stroke: var(--line); stroke-dasharray: 2 3; }}
    .line-axis {{ stroke: var(--line-2); }}
    .line-path {{
      fill: none;
      stroke: var(--accent);
      stroke-width: 2;
      stroke-linejoin: round;
    }}
    .line-area {{ fill: rgba(154, 43, 39, 0.07); }}
    .line-point {{
      fill: var(--panel);
      stroke: var(--accent);
      stroke-width: 2;
    }}
    .line-label, .line-value {{
      fill: var(--muted);
      font-family: var(--mono);
      font-size: 11px;
    }}
    .line-value {{ fill: var(--ink); font-weight: 600; }}

    table {{ width: 100%; border-collapse: collapse; }}

    th, td {{
      padding: 10px 8px;
      border-bottom: 1px solid var(--line);
      text-align: left;
      vertical-align: top;
      font-size: 13px;
      line-height: 1.4;
    }}

    th {{
      color: var(--ink);
      background: var(--paper-2);
      font-size: 10.5px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      border-bottom: 2px solid var(--rule);
    }}

    .scroll-table {{
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 3px;
      background: var(--panel);
    }}

    .scroll-table table {{ min-width: 720px; }}

    .scroll-table .master-table {{ min-width: 3260px; table-layout: fixed; }}
    .master-table th, .master-table td {{ white-space: nowrap; }}
    .master-table td.numeric {{ text-align: right; font-family: var(--mono); }}
    .master-table .master-groups th {{
      color: var(--panel);
      background: var(--ink);
      border-right: 1px solid var(--ink-soft);
      text-align: center;
      letter-spacing: 0.07em;
    }}
    .master-table .master-groups th:first-child {{ position: sticky; left: 0; z-index: 4; }}
    .master-table col:nth-child(1) {{ width: 58px; }}
    .master-table col:nth-child(2) {{ width: 106px; }}
    .master-table col:nth-child(3) {{ width: 144px; }}
    .master-table col:nth-child(4) {{ width: 230px; }}
    .master-table col:nth-child(5) {{ width: 310px; }}
    .master-table col:nth-child(n+6):nth-child(-n+24) {{ width: 98px; }}
    .master-table th:nth-child(5), .master-table td:nth-child(5) {{ white-space: normal; }}
    .master-table th:nth-child(9), .master-table td:nth-child(9), .master-table th:nth-child(10), .master-table td:nth-child(10) {{ text-align: center; }}
    .master-table col:nth-child(25) {{ width: 470px; }}
    .master-table th:nth-child(25), .master-table td:nth-child(25) {{ white-space: normal; }}
    .master-table thead tr:last-child th:nth-child(-n+3), .master-table tbody td:nth-child(-n+3) {{ position: sticky; z-index: 2; background: var(--panel); }}
    .master-table thead tr:last-child th:nth-child(-n+3) {{ z-index: 3; background: var(--paper-2); }}
    .master-table th:nth-child(1), .master-table td:nth-child(1) {{ left: 0; }}
    .master-table th:nth-child(2), .master-table td:nth-child(2) {{ left: 58px; }}
    .master-table th:nth-child(3), .master-table td:nth-child(3) {{ left: 164px; box-shadow: 6px 0 8px -8px rgba(27, 26, 22, 0.65); }}

    #heatmap th {{ text-transform: uppercase; font-size: 10.5px; line-height: 1.3; }}
    #heatmap td {{ font-family: var(--mono); }}

    .speech-list {{ display: grid; gap: 10px; }}

    .speech-card {{
      padding: 14px;
      border: 1px solid var(--line);
      border-left: 3px solid var(--line-2);
      border-radius: 3px;
      background: var(--panel);
    }}

    .speech-card strong {{
      display: block;
      margin-bottom: 6px;
      line-height: 1.35;
      font-size: 14px;
    }}

    .speech-meta {{ color: var(--muted); font-size: 12px; line-height: 1.45; }}

    .pager {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 12px; }}

    .pager button {{
      min-height: 46px;
      border: 1px solid var(--rule);
      border-radius: 2px;
      background: transparent;
      color: var(--ink);
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      font-size: 13px;
    }}

    .pager button:disabled {{ opacity: 0.4; }}

    .framework-item {{
      padding: 16px;
      border: 1px solid var(--line);
      border-left: 3px solid var(--accent);
      border-radius: 3px;
      background: var(--panel);
      margin-bottom: 12px;
    }}

    .framework-item h3 {{
      margin: 0 0 8px;
      color: var(--accent);
      font-family: var(--serif);
      font-size: 19px;
      font-weight: 600;
    }}

    .framework-item p {{ margin: 0 0 6px; font-size: 13px; line-height: 1.5; color: var(--muted); }}

    .research-switch {{ margin-bottom: 14px; }}

    .mobile-rate-row {{
      margin: 14px 0;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
    }}

    .mobile-rate-head {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      margin-bottom: 7px;
      font-size: 13px;
    }}

    .mobile-rate-head span {{ color: var(--muted); font-family: var(--mono); white-space: nowrap; }}
    .mobile-stack {{ display: flex; height: 13px; overflow: hidden; background: var(--paper-2); }}
    .mobile-stack i {{ display: block; height: 100%; }}

    .mobile-finding {{
      padding: 12px 0;
      border-bottom: 1px solid var(--line);
      color: var(--ink-soft);
      font-size: 13px;
      line-height: 1.45;
    }}

    .mobile-finding strong {{ color: var(--accent); font-family: var(--serif); font-size: 21px; margin-right: 7px; }}

    @media (min-width: 760px) {{
      body::before {{
        content: "This is the mobile dashboard. Open ../interactive_dashboard/index.html for the desktop version.";
        display: block;
        padding: 10px 14px;
        color: var(--paper);
        background: var(--ink);
        text-align: center;
        font-size: 13px;
      }}
    }}
  
</style>
</head>
<body>
  <header>
    <p class="kicker">Mobile Dashboard</p>
    <h1>Prabowo Speeches</h1>
    <p class="subtitle">A phone-friendly view of situational context and linguistic patterns across the speech corpus.</p>
  </header>

  <nav class="mobile-nav" aria-label="Mobile dashboard sections">
    <button class="active" data-view="overview">Overview</button>
    <button data-view="summary">Summary</button>
    <button data-view="framework">Framework</button>
    <button data-view="table">Table</button>
    <button data-view="research">Analysis</button>
  </nav>

  <main>
    <section id="overview" class="view active">
      <div class="mobile-filters" data-filter-surface>
        <input id="searchInput" type="search" placeholder="Search speeches" />
        <select id="languageFilter"></select>
        <select id="yearFilter"></select>
        <select id="interactionFilter"></select>
        <select id="domainFilter"></select>
        <button class="reset" id="clearFilters">Reset Filters</button>
      </div>
      <div class="kpis" id="kpis"></div>
      <article class="card">
        <h2>Language Mix</h2>
        <div id="languageChart"></div>
      </article>
      <article class="card">
        <h2>Monthly Speech Trend</h2>
        <div id="timelineChart"></div>
      </article>
      <article class="card">
        <h2>Interaction Type</h2>
        <div id="interactionChart"></div>
      </article>
      <article class="card">
        <h2>Field Domain</h2>
        <div id="domainChart"></div>
      </article>
      <article class="card">
        <h2>Interaction by Domain</h2>
        <div class="scroll-table" id="heatmap"></div>
      </article>
    </section>

    <section id="summary" class="view">
      <article class="card">
        <h2>Corpus and Languages</h2>
        <div id="summaryCorpus"></div>
      </article>
      <article class="card">
        <h2>Interaction Type</h2>
        <div id="summaryInteraction"></div>
      </article>
      <article class="card">
        <h2>Field Domain</h2>
        <div id="summaryDomain"></div>
      </article>
      <article class="card">
        <h2>Visual Mapping Counts</h2>
        <div id="visualMapping"></div>
      </article>
    </section>

    <section id="framework" class="view">
      <article class="card">
        <h2>Framework</h2>
        <div id="frameworkList"></div>
      </article>
    </section>

    <section id="table" class="view">
      <div class="mobile-filters" data-filter-surface>
        <input id="tableSearchInput" type="search" placeholder="Search speeches" />
        <select id="tableLanguageFilter"></select>
        <select id="tableYearFilter"></select>
        <select id="tableInteractionFilter"></select>
        <select id="tableDomainFilter"></select>
        <button class="reset" id="tableClearFilters">Reset Filters</button>
      </div>
      <article class="card">
        <h2>Analysis Table</h2>
        <p class="speech-meta" id="tableStatus"></p>
        <div class="speech-list" id="speechList"></div>
        <div class="pager">
          <button id="prevPage">Previous</button>
          <button id="nextPage">Next</button>
        </div>
      </article>
    </section>

    <section id="research" class="view">
      <article class="card">
        <h2>Linguistic Analysis</h2>
        <p class="speech-meta">Domestic sample: 29 speeches · International sample: 13 speeches. Rates are per 1,000 presidential words.</p>
        <select id="mobileResearchSelect" class="research-switch" aria-label="Linguistic analysis view">
          <option value="selfmention">Self-Mention</option>
          <option value="appraisal">Appraisal</option>
          <option value="engagement">Engagement</option>
          <option value="frames">Verb Frames</option>
          <option value="quality">Data Quality</option>
          <option value="sources">Source Tables</option>
        </select>
        <div id="mobileResearchContent"></div>
      </article>
    </section>
  </main>

  <script>
    const DATA = {data_json};
    const state = {{
      view: "overview",
      search: "",
      language: "All languages",
      year: "All years",
      interaction: "All interaction types",
      domain: "All field domains",
      page: 1,
      pageSize: 12,
      researchView: "selfmention",
      researchTable: "T0_per_speech_master"
    }};
    const colors = ["#9a2b27", "#2f4858", "#876c34", "#4c5d4d", "#5a4a5c"];
    const fmt = new Intl.NumberFormat("en-US");
    const pct = value => `${{Math.round((Number(value) || 0) * 1000) / 10}}%`;
    const esc = value => String(value ?? "").replace(/[&<>"']/g, char => ({{ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }}[char]));
    const titleCase = text => String(text || "").replace(/\\w\\S*/g, word => word.charAt(0).toUpperCase() + word.slice(1).toLowerCase());

    function uniq(values) {{
      return [...new Set(values.filter(Boolean))].sort((a, b) => String(a).localeCompare(String(b)));
    }}

    function countBy(rows, key) {{
      const map = new Map();
      rows.forEach(row => {{
        const value = row[key] || "Unspecified";
        map.set(value, (map.get(value) || 0) + 1);
      }});
      return [...map.entries()].map(([label, value]) => ({{ label, value }})).sort((a, b) => b.value - a.value || a.label.localeCompare(b.label));
    }}

    function filteredRows() {{
      return DATA.analysis.filter(row => {{
        const text = [row.event, row.interactionType, row.fieldDomain, row.field, row.tenor, row.mode, row.openingEvidence].join(" ").toLowerCase();
        return (!state.search || text.includes(state.search))
          && (state.language === "All languages" || row.language === state.language)
          && (state.year === "All years" || row.year === state.year)
          && (state.interaction === "All interaction types" || row.interactionType === state.interaction)
          && (state.domain === "All field domains" || row.fieldDomain === state.domain);
      }});
    }}

    function fillSelect(id, label, values) {{
      const el = document.getElementById(id);
      el.innerHTML = [`<option>${{label}}</option>`, ...values.map(value => `<option>${{esc(value)}}</option>`)].join("");
      el.value = stateFromLabel(label);
    }}

    function stateFromLabel(label) {{
      if (label.includes("languages")) return state.language;
      if (label.includes("years")) return state.year;
      if (label.includes("interaction")) return state.interaction;
      if (label.includes("field")) return state.domain;
      return label;
    }}

    function syncFilterInputs() {{
      const pairs = [
        ["searchInput", state.search], ["tableSearchInput", state.search],
        ["languageFilter", state.language], ["tableLanguageFilter", state.language],
        ["yearFilter", state.year], ["tableYearFilter", state.year],
        ["interactionFilter", state.interaction], ["tableInteractionFilter", state.interaction],
        ["domainFilter", state.domain], ["tableDomainFilter", state.domain]
      ];
      pairs.forEach(([id, value]) => {{
        const el = document.getElementById(id);
        if (el) el.value = value;
      }});
    }}

    function initFilters() {{
      ["languageFilter", "tableLanguageFilter"].forEach(id => fillSelect(id, "All languages", uniq(DATA.analysis.map(d => d.language))));
      ["yearFilter", "tableYearFilter"].forEach(id => fillSelect(id, "All years", uniq(DATA.analysis.map(d => d.year))));
      ["interactionFilter", "tableInteractionFilter"].forEach(id => fillSelect(id, "All interaction types", uniq(DATA.analysis.map(d => d.interactionType))));
      ["domainFilter", "tableDomainFilter"].forEach(id => fillSelect(id, "All field domains", uniq(DATA.analysis.map(d => d.fieldDomain))));

      [["searchInput", "search"], ["tableSearchInput", "search"]].forEach(([id]) => {{
        document.getElementById(id).addEventListener("input", event => {{
          state.search = event.target.value.trim().toLowerCase();
          state.page = 1;
          syncFilterInputs();
          render();
        }});
      }});
      [["languageFilter", "language"], ["tableLanguageFilter", "language"], ["yearFilter", "year"], ["tableYearFilter", "year"], ["interactionFilter", "interaction"], ["tableInteractionFilter", "interaction"], ["domainFilter", "domain"], ["tableDomainFilter", "domain"]].forEach(([id, key]) => {{
        document.getElementById(id).addEventListener("change", event => {{
          state[key] = event.target.value;
          state.page = 1;
          syncFilterInputs();
          render();
        }});
      }});
      ["clearFilters", "tableClearFilters"].forEach(id => document.getElementById(id).addEventListener("click", () => {{
        state.search = "";
        state.language = "All languages";
        state.year = "All years";
        state.interaction = "All interaction types";
        state.domain = "All field domains";
        state.page = 1;
        syncFilterInputs();
        render();
      }}));
    }}

    function renderKpis(rows) {{
      const words = rows.reduce((sum, row) => sum + (Number(row.wordCount) || 0), 0);
      const topInteraction = countBy(rows, "interactionType")[0];
      const topDomain = countBy(rows, "fieldDomain")[0];
      const kpis = [
        ["Speeches", fmt.format(rows.length), `${{fmt.format(DATA.analysis.length)}} total`],
        ["Words", fmt.format(words), `${{fmt.format(Math.round(words / Math.max(rows.length, 1)))}} avg`],
        ["Languages", fmt.format(uniq(rows.map(row => row.language)).length), countBy(rows, "language").map(d => `${{d.label}} ${{d.value}}`).join(" | ")],
        ["Top Type", topInteraction ? fmt.format(topInteraction.value) : "0", topInteraction?.label || "No match"],
        ["Top Domain", topDomain ? fmt.format(topDomain.value) : "0", topDomain?.label || "No match"]
      ];
      document.getElementById("kpis").innerHTML = kpis.map(kpi => `
        <div class="kpi"><span>${{esc(kpi[0])}}</span><strong>${{esc(kpi[1])}}</strong><small>${{esc(kpi[2])}}</small></div>
      `).join("");
    }}

    function renderBars(id, rows, key, className = "") {{
      const data = countBy(rows, key);
      const max = Math.max(...data.map(d => d.value), 1);
      document.getElementById(id).innerHTML = data.map(d => `
        <div class="bar-row">
          <div class="bar-label">${{esc(d.label)}}</div>
          <div class="bar-value">${{fmt.format(d.value)}} (${{pct(d.value / Math.max(rows.length, 1))}})</div>
          <div class="bar-track"><div class="bar-fill ${{className}}" style="width:${{Math.max(3, d.value / max * 100)}}%"></div></div>
        </div>
      `).join("") || `<p class="speech-meta">No matching speeches.</p>`;
    }}

    function renderLanguage(rows) {{
      const data = countBy(rows, "language");
      const total = rows.length || 1;
      let angle = 0;
      const stops = data.map((d, idx) => {{
        const start = angle;
        angle += (d.value / total) * 360;
        return `${{colors[idx % colors.length]}} ${{start}}deg ${{angle}}deg`;
      }}).join(", ");
      document.getElementById("languageChart").innerHTML = `
        <div class="donut-wrap">
          <div class="donut" style="background:${{stops ? `conic-gradient(${{stops}})` : "#f0e6df"}}"></div>
          <div class="legend">
            ${{data.map((d, idx) => `<div class="legend-row"><span class="swatch" style="background:${{colors[idx % colors.length]}}"></span><span>${{esc(d.label)}}</span><strong>${{d.value}} (${{pct(d.value / total)}})</strong></div>`).join("")}}
          </div>
        </div>
      `;
    }}

    function renderTimeline(rows) {{
      const map = new Map();
      rows.forEach(row => {{
        if (!row.date) return;
        const key = row.date.slice(0, 7);
        map.set(key, (map.get(key) || 0) + 1);
      }});
      const data = [...map.entries()].sort().map(([label, value]) => ({{ label, value }}));
      if (!data.length) {{
        document.getElementById("timelineChart").innerHTML = `<p class="speech-meta">No matching dates.</p>`;
        return;
      }}
      const width = 360;
      const height = 210;
      const pad = {{ left: 26, right: 12, top: 24, bottom: 34 }};
      const max = Math.max(...data.map(d => d.value), 1);
      const plotWidth = width - pad.left - pad.right;
      const plotHeight = height - pad.top - pad.bottom;
      const bottomY = height - pad.bottom;
      const points = data.map((d, index) => {{
        const x = data.length === 1 ? pad.left + plotWidth / 2 : pad.left + (index / (data.length - 1)) * plotWidth;
        const y = bottomY - (d.value / max) * plotHeight;
        return {{ ...d, x, y }};
      }});
      const path = points.map((point, index) => `${{index ? "L" : "M"}} ${{point.x.toFixed(1)}} ${{point.y.toFixed(1)}}`).join(" ");
      const area = `${{path}} L ${{points[points.length - 1].x.toFixed(1)}} ${{bottomY}} L ${{points[0].x.toFixed(1)}} ${{bottomY}} Z`;
      const labelEvery = Math.max(1, Math.ceil(data.length / 5));
      document.getElementById("timelineChart").innerHTML = `
        <svg class="line-chart" viewBox="0 0 ${{width}} ${{height}}" role="img" aria-label="Monthly speech trend">
          <line class="line-grid" x1="${{pad.left}}" x2="${{width - pad.right}}" y1="${{pad.top}}" y2="${{pad.top}}"></line>
          <line class="line-grid" x1="${{pad.left}}" x2="${{width - pad.right}}" y1="${{bottomY - plotHeight / 2}}" y2="${{bottomY - plotHeight / 2}}"></line>
          <line class="line-axis" x1="${{pad.left}}" x2="${{width - pad.right}}" y1="${{bottomY}}" y2="${{bottomY}}"></line>
          <path class="line-area" d="${{area}}"></path>
          <path class="line-path" d="${{path}}"></path>
          ${{points.map(point => `<circle class="line-point" cx="${{point.x.toFixed(1)}}" cy="${{point.y.toFixed(1)}}" r="4"></circle>`).join("")}}
          ${{points.map(point => `<text class="line-value" x="${{point.x.toFixed(1)}}" y="${{Math.max(13, point.y - 9).toFixed(1)}}" text-anchor="middle">${{point.value}}</text>`).join("")}}
          ${{points.map((point, index) => index % labelEvery === 0 || index === points.length - 1 ? `<text class="line-label" x="${{point.x.toFixed(1)}}" y="${{height - 10}}" text-anchor="middle">${{esc(point.label.slice(5) + "/" + point.label.slice(2, 4))}}</text>` : "").join("")}}
        </svg>
      `;
    }}

    function renderHeatmap(rows) {{
      const interactions = countBy(rows, "interactionType").map(d => d.label);
      const domains = countBy(rows, "fieldDomain").map(d => d.label);
      const max = Math.max(...interactions.flatMap(i => domains.map(d => rows.filter(row => row.interactionType === i && row.fieldDomain === d).length)), 1);
      const head = `<tr><th>Interaction</th>${{domains.map(d => `<th>${{esc(titleCase(d))}}</th>`).join("")}}</tr>`;
      const body = interactions.map(interaction => `
        <tr>
          <td><strong>${{esc(titleCase(interaction))}}</strong></td>
          ${{domains.map(domain => {{
            const value = rows.filter(row => row.interactionType === interaction && row.fieldDomain === domain).length;
            const alpha = value ? 0.12 + (value / max) * 0.62 : 0;
            return `<td style="background:rgba(140,42,39,${{alpha}}); color:${{alpha > 0.55 ? "#f4f1ea" : "inherit"}}; text-align:center; font-variant-numeric:tabular-nums;">${{value || ""}}</td>`;
          }}).join("")}}
        </tr>
      `).join("");
      document.getElementById("heatmap").innerHTML = rows.length ? `<table>${{head}}${{body}}</table>` : `<p class="speech-meta">No matching speeches.</p>`;
    }}

    function summaryTable(rows) {{
      return `<table><thead><tr><th>Category</th><th>Count</th><th>Share</th></tr></thead><tbody>${{rows.map(row => `<tr><td>${{esc(row.category)}}</td><td>${{fmt.format(row.count || 0)}}</td><td>${{typeof row.share === "number" ? pct(row.share) : esc(row.share || "")}}</td></tr>`).join("")}}</tbody></table>`;
    }}

    function renderSummary() {{
      document.getElementById("summaryCorpus").innerHTML = summaryTable(DATA.summary.filter(row => row.metric === "Total speeches" || row.metric === "Language group"));
      document.getElementById("summaryInteraction").innerHTML = summaryTable(DATA.summary.filter(row => row.metric === "Interaction type"));
      document.getElementById("summaryDomain").innerHTML = summaryTable(DATA.summary.filter(row => row.metric === "Field domain"));
      document.getElementById("visualMapping").innerHTML = `
        <h3>Interaction Types</h3>
        ${{DATA.visual.interactionTypes.map(item => `<div class="bar-row"><div class="bar-label">${{esc(item.category)}}</div><div class="bar-value">${{item.count}} (${{pct(item.share)}})</div><div class="bar-track"><div class="bar-fill" style="width:${{Math.max(3, item.share * 100)}}%"></div></div></div>`).join("")}}
        <h3>Field Domains</h3>
        ${{DATA.visual.fieldDomains.map(item => `<div class="bar-row"><div class="bar-label">${{esc(item.category)}}</div><div class="bar-value">${{item.count}} (${{pct(item.share)}})</div><div class="bar-track"><div class="bar-fill indigo" style="width:${{Math.max(3, item.share * 100)}}%"></div></div></div>`).join("")}}
      `;
    }}

    function renderFramework() {{
      document.getElementById("frameworkList").innerHTML = DATA.framework.map(item => `
        <div class="framework-item">
          <h3>${{esc(item.term)}}</h3>
          <p>${{esc(item.meaning || "")}}</p>
          <p class="speech-meta">${{esc(item.components || "")}}</p>
        </div>
      `).join("");
    }}

    function renderSpeechList(rows) {{
      const totalPages = Math.max(1, Math.ceil(rows.length / state.pageSize));
      state.page = Math.min(state.page, totalPages);
      const start = (state.page - 1) * state.pageSize;
      const pageRows = rows.slice(start, start + state.pageSize);
      document.getElementById("tableStatus").textContent = `${{fmt.format(rows.length)}} matching speeches | Page ${{state.page}} of ${{totalPages}}`;
      document.getElementById("speechList").innerHTML = pageRows.map(row => `
        <article class="speech-card">
          <strong>${{esc(row.no)}}. ${{esc(row.event)}}</strong>
          <div class="speech-meta">${{esc(row.date)}} | ${{esc(row.language)}} | ${{esc(row.interactionType)}} | ${{fmt.format(row.wordCount)}} words</div>
          <div class="speech-meta">${{esc(row.fieldDomain)}}</div>
        </article>
      `).join("") || `<p class="speech-meta">No matching speeches.</p>`;
      document.getElementById("prevPage").disabled = state.page <= 1;
      document.getElementById("nextPage").disabled = state.page >= totalPages;
    }}

    function mobileResearchRows(key) {{
      return DATA.research?.tables?.[key]?.rows || [];
    }}

    function mobileSourceCell(value) {{
      if (value === true) return "Yes";
      if (value === false) return "No";
      if (value === "" || value == null) return "—";
      return value;
    }}

    function mobileRateRows(rows, valueKey = "per_1000w", color = "var(--accent)") {{
      const max = Math.max(...rows.map(row => Number(row[valueKey]) || 0), 1);
      return rows.map(row => `
        <div class="mobile-rate-row">
          <div class="mobile-rate-head"><strong>${{esc(row.form || row.frame || row.pronoun_node)}}</strong><span>${{Number(row[valueKey]).toLocaleString("en-US", {{ maximumFractionDigits: 2 }})}} / 1k</span></div>
          <div class="bar-track"><div class="bar-fill" style="background:${{color}};width:${{Math.max(2, Number(row[valueKey]) / max * 100)}}%"></div></div>
          ${{row.function ? `<p class="speech-meta">${{esc(row.function)}} · ${{fmt.format(Number(row.raw) || 0)}} raw</p>` : ""}}
        </div>`).join("");
    }}

    function renderMobileResearch() {{
      const target = document.getElementById("mobileResearchContent");
      if (state.researchView === "selfmention") {{
        const rows = mobileResearchRows("T2_selfmention_sample").filter(row => row.form !== "TOTAL");
        target.innerHTML = `<h3>Domestic Sample</h3>${{mobileRateRows(rows.filter(row => String(row.sample).startsWith("Domestic")))}}<h3>International Sample</h3>${{mobileRateRows(rows.filter(row => String(row.sample).startsWith("International")), "per_1000w", "var(--slate)")}}`;
      }} else if (state.researchView === "appraisal") {{
        const rows = mobileResearchRows("T5b_attitude_group_summary");
        target.innerHTML = `<p class="speech-meta">Affect / Judgement / Appreciation share around each pronoun node.</p>${{rows.map(row => `
          <div class="mobile-rate-row">
            <div class="mobile-rate-head"><strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}}</strong><span>${{row.attitude_per_1000w}} / 1k</span></div>
            <div class="mobile-stack"><i style="width:${{row.AFF_pct}}%;background:var(--accent)"></i><i style="width:${{row.JUD_pct}}%;background:var(--slate)"></i><i style="width:${{row.APP_pct}}%;background:var(--ochre)"></i></div>
            <p class="speech-meta">Affect ${{row.AFF_pct}}% · Judgement ${{row.JUD_pct}}% · Appreciation ${{row.APP_pct}}%</p>
          </div>`).join("")}}`;
      }} else if (state.researchView === "engagement") {{
        const rows = mobileResearchRows("T6_engagement_hedge_booster");
        const max = Math.max(...rows.flatMap(row => [Number(row.hedges_per1k), Number(row.boosters_per1k)]), 1);
        target.innerHTML = `<p class="speech-meta">Ochre: hedges · Slate: boosters</p>${{rows.map(row => `
          <div class="mobile-rate-row">
            <div class="mobile-rate-head"><strong>${{esc(row.sample)}} · ${{esc(row.pronoun_node)}}</strong><span>ratio ${{row.hedge_booster_ratio}}</span></div>
            <div class="bar-track"><div class="bar-fill" style="background:var(--ochre);width:${{Math.max(2, Number(row.hedges_per1k) / max * 100)}}%"></div></div>
            <div class="bar-track" style="margin-top:4px"><div class="bar-fill" style="background:var(--slate);width:${{Math.max(2, Number(row.boosters_per1k) / max * 100)}}%"></div></div>
            <p class="speech-meta">${{row.hedges_per1k}} hedges · ${{row.boosters_per1k}} boosters per 1,000</p>
          </div>`).join("")}}`;
      }} else if (state.researchView === "frames") {{
        const rows = mobileResearchRows("T7_firstperson_verb_frames");
        const sampleRows = sample => [...rows.filter(row => String(row.sample).startsWith(sample))].sort((a, b) => Number(b.per_1000w) - Number(a.per_1000w));
        const byFunction = items => {{
          const groups = new Map();
          items.forEach(row => {{
            const current = groups.get(row.function) || {{ frame: row.function, raw: 0, per_1000w: 0 }};
            current.raw += Number(row.raw) || 0;
            current.per_1000w += Number(row.per_1000w) || 0;
            groups.set(row.function, current);
          }});
          return [...groups.values()].sort((a, b) => b.per_1000w - a.per_1000w);
        }};
        const domestic = sampleRows("Domestic");
        const international = sampleRows("International");
        target.innerHTML = `<h3>Domestic Functional Profile</h3>${{mobileRateRows(byFunction(domestic))}}<h3>International Functional Profile</h3>${{mobileRateRows(byFunction(international), "per_1000w", "var(--slate)")}}<h3>Domestic Verb Frames</h3>${{mobileRateRows(domestic)}}<h3>International Verb Frames</h3>${{mobileRateRows(international, "per_1000w", "var(--slate)")}}`;
      }} else if (state.researchView === "quality") {{
        const rows = mobileResearchRows("T8_data_hygiene_log").filter(row => Number(row.words_removed) > 0);
        const removed = rows.reduce((sum, row) => sum + Number(row.words_removed), 0);
        const exclusions = Object.entries(DATA.research.samples.excluded || {{}});
        target.innerHTML = `<div class="mobile-finding"><strong>${{rows.length}}</strong>multi-speaker transcripts cleaned.</div><div class="mobile-finding"><strong>${{fmt.format(removed)}}</strong>non-Prabowo words removed.</div><h3>Sample Exclusions</h3>${{exclusions.map(([id, reason]) => `<div class="mobile-finding"><strong>${{id}}</strong>${{esc(reason)}}</div>`).join("")}}`;
      }} else {{
        const tables = DATA.research.tables;
        const keys = Object.keys(tables);
        const table = tables[state.researchTable] || tables[keys[0]];
        const headers = table.rows.length ? Object.keys(table.rows[0]) : [];
        const isMaster = state.researchTable === "T0_per_speech_master";
        const masterGroups = isMaster ? `<tr class="master-groups"><th colspan="3">Identity</th><th colspan="2">Discourse Context</th><th colspan="3">Corpus Words</th><th colspan="2">Sample Membership</th><th colspan="9">Raw Self-Mentions</th><th colspan="5">Rate per 1,000</th><th>Speech</th></tr>` : "";
        const masterCols = isMaster ? `<colgroup>${{headers.map(() => "<col>").join("")}}</colgroup>` : "";
        target.innerHTML = `<select id="mobileResearchTable" aria-label="Source table">${{keys.map(key => `<option value="${{esc(key)}}" ${{tables[key] === table ? "selected" : ""}}>${{esc(tables[key].title)}}</option>`).join("")}}</select><p class="speech-meta">${{esc(table.source)}} · ${{table.rows.length}} rows</p><div class="scroll-table"><table class="${{isMaster ? "master-table" : ""}}">${{masterCols}}<thead>${{masterGroups}}<tr>${{headers.map(header => `<th>${{esc(header.replaceAll("_", " "))}}</th>`).join("")}}</tr></thead><tbody>${{table.rows.map(row => `<tr>${{headers.map(header => `<td class="${{typeof row[header] === "number" ? "numeric" : ""}}">${{esc(mobileSourceCell(row[header]))}}</td>`).join("")}}</tr>`).join("")}}</tbody></table></div>`;
        document.getElementById("mobileResearchTable").addEventListener("change", event => {{ state.researchTable = event.target.value; renderMobileResearch(); }});
      }}
    }}

    function bindNav() {{
      document.querySelectorAll(".mobile-nav button").forEach(button => button.addEventListener("click", () => {{
        state.view = button.dataset.view;
        document.querySelectorAll(".mobile-nav button").forEach(item => item.classList.toggle("active", item === button));
        document.querySelectorAll(".view").forEach(view => view.classList.toggle("active", view.id === state.view));
      }}));
      document.getElementById("prevPage").addEventListener("click", () => {{
        state.page = Math.max(1, state.page - 1);
        render();
      }});
      document.getElementById("nextPage").addEventListener("click", () => {{
        state.page += 1;
        render();
      }});
      document.getElementById("mobileResearchSelect").addEventListener("change", event => {{
        state.researchView = event.target.value;
        renderMobileResearch();
      }});
    }}

    function render() {{
      const rows = filteredRows();
      renderKpis(rows);
      renderLanguage(rows);
      renderTimeline(rows);
      renderBars("interactionChart", rows, "interactionType");
      renderBars("domainChart", rows, "fieldDomain", "indigo");
      renderHeatmap(rows);
      renderSummary();
      renderFramework();
      renderSpeechList(rows);
      renderMobileResearch();
    }}

    initFilters();
    bindNav();
    syncFilterInputs();
    render();
  </script>
</body>
</html>
"""


def main() -> None:
    analysis_rows = normalize_analysis(sheet_rows(DISCOURSE_XLSX, "Analysis Table", header_row=1))
    data = {
        "analysis": analysis_rows,
        "framework": framework_rows(),
        "summary": summary_rows(),
        "visual": visual_mapping(),
        "research": research_analysis(),
        "meta": {
            "generatedFrom": [DISCOURSE_XLSX.name, VISUAL_XLSX.name],
            "speechCount": len(analysis_rows),
        },
    }
    OUT_DIR.mkdir(exist_ok=True)
    OUT_FILE.write_text(build_html(data), encoding="utf-8")
    MOBILE_OUT_DIR.mkdir(exist_ok=True)
    MOBILE_OUT_FILE.write_text(build_mobile_html(data), encoding="utf-8")
    print(OUT_FILE)
    print(MOBILE_OUT_FILE)


if __name__ == "__main__":
    main()
